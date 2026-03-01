from __future__ import annotations

import csv
import io
from datetime import timedelta
from typing import Callable, Dict, List, Tuple

from django.contrib.auth import get_user_model
from django.db.models import Avg, Count, F, Q
from django.utils import timezone
from openpyxl import Workbook

from sims.academics.models import Department
from sims.audit.models import ActivityLog
from sims.bulk.models import BulkOperation
from sims.cases.models import ClinicalCase
from sims.logbook.models import LogbookEntry
from sims.rotations.models import HospitalDepartment, Rotation

User = get_user_model()

Rows = List[Dict[str, object]]
Summary = Dict[str, object]
Runner = Callable[[object, dict], Tuple[Rows, Summary]]


def _pg_scope(user):
    if user.role in {"admin", "utrmc_admin", "utrmc_user"}:
        return User.objects.filter(role="pg")
    if user.role == "supervisor":
        return User.objects.filter(role="pg", supervisor=user)
    return User.objects.filter(id=user.id, role="pg")


def _apply_pg_filters(queryset, filters: dict):
    department_id = filters.get("department_id")
    training_year = filters.get("year")
    if department_id:
        queryset = queryset.filter(home_department_id=department_id)
    if training_year:
        queryset = queryset.filter(year=str(training_year))
    return queryset


def _residents_roster(user, filters):
    rows = []
    for pg in _apply_pg_filters(_pg_scope(user).select_related("supervisor", "home_department"), filters):
        rows.append(
            {
                "username": pg.username,
                "name": pg.get_full_name(),
                "year": pg.year or "",
                "specialty": pg.specialty or "",
                "department": pg.home_department.name if pg.home_department else "",
                "supervisor": pg.supervisor.get_full_name() if pg.supervisor else "",
            }
        )
    return rows, {"total": len(rows)}


def _supervisors_roster(user, _filters):
    qs = User.objects.filter(role="supervisor").annotate(assigned_count=Count("assigned_pgs"))
    rows = [
        {
            "username": sup.username,
            "name": sup.get_full_name(),
            "specialty": sup.specialty or "",
            "assigned_pgs": sup.assigned_count,
        }
        for sup in qs
    ]
    return rows, {"total": len(rows)}


def _departments_mapping(_user, _filters):
    rows = []
    for dept in Department.objects.all().order_by("name"):
        mapped = HospitalDepartment.objects.filter(department=dept, is_active=True).count()
        rows.append(
            {
                "code": dept.code,
                "name": dept.name,
                "active": dept.active,
                "mapped_hospitals": mapped,
                "coverage": "mapped" if mapped else "unmapped",
            }
        )
    uncovered = len([r for r in rows if r["mapped_hospitals"] == 0])
    return rows, {"total": len(rows), "unmapped_departments": uncovered}


def _residents_on_rotation(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = [
        {
            "resident": r.pg.get_full_name(),
            "department": r.department.name,
            "hospital": r.hospital.name,
            "start_date": r.start_date.isoformat(),
            "end_date": r.end_date.isoformat(),
            "status": r.status,
        }
        for r in Rotation.objects.select_related("pg", "department", "hospital")
        .filter(pg__in=pgs, status="ongoing")
        .order_by("pg__last_name")
    ]
    return rows, {"total": len(rows)}


def _residents_without_active_rotation(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    active_ids = Rotation.objects.filter(status="ongoing").values_list("pg_id", flat=True)
    qs = pgs.exclude(id__in=active_ids)
    rows = [
        {"resident": pg.get_full_name(), "username": pg.username, "department": pg.home_department.name if pg.home_department else ""}
        for pg in qs.select_related("home_department")
    ]
    return rows, {"total": len(rows)}


def _upcoming_rotations(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = [
        {
            "resident": r.pg.get_full_name(),
            "department": r.department.name,
            "hospital": r.hospital.name,
            "start_date": r.start_date.isoformat(),
            "end_date": r.end_date.isoformat(),
        }
        for r in Rotation.objects.select_related("pg", "department", "hospital")
        .filter(pg__in=pgs, start_date__gt=timezone.now().date())
        .order_by("start_date")[:200]
    ]
    return rows, {"total": len(rows)}


def _rotation_coverage(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = list(
        Rotation.objects.filter(pg__in=pgs)
        .values(department_name=F("department__name"))
        .annotate(total=Count("id"), ongoing=Count("id", filter=Q(status="ongoing")))
        .order_by("department_name")
    )
    return rows, {"departments": len(rows)}


def _logbook_compliance(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = []
    for pg in pgs:
        entries = LogbookEntry.objects.filter(pg=pg)
        total = entries.count()
        approved = entries.filter(status="approved").count()
        rows.append(
            {
                "resident": pg.get_full_name(),
                "total_entries": total,
                "approved_entries": approved,
                "compliance_pct": round((approved / total) * 100, 2) if total else 0.0,
            }
        )
    return rows, {"total": len(rows)}


def _logbook_uptodate(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    cutoff = timezone.now().date() - timedelta(days=30)
    rows = []
    for pg in pgs:
        recent = LogbookEntry.objects.filter(pg=pg, date__gte=cutoff).count()
        rows.append({"resident": pg.get_full_name(), "entries_last_30_days": recent, "up_to_date": recent > 0})
    return rows, {"total": len(rows)}


def _pending_logbook_queue(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = [
        {
            "id": entry.id,
            "resident": entry.pg.get_full_name(),
            "case_title": entry.case_title,
            "submitted_at": entry.submitted_to_supervisor_at.isoformat() if entry.submitted_to_supervisor_at else "",
            "supervisor": entry.supervisor.get_full_name() if entry.supervisor else "",
        }
        for entry in LogbookEntry.objects.select_related("pg", "supervisor")
        .filter(pg__in=pgs, status="pending")
        .order_by("-submitted_to_supervisor_at")
    ]
    return rows, {"total": len(rows)}


def _overdue_verification(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    threshold = timezone.now() - timedelta(days=7)
    rows = [
        {
            "id": entry.id,
            "resident": entry.pg.get_full_name(),
            "supervisor": entry.supervisor.get_full_name() if entry.supervisor else "",
            "submitted_at": entry.submitted_to_supervisor_at.isoformat() if entry.submitted_to_supervisor_at else "",
        }
        for entry in LogbookEntry.objects.select_related("pg", "supervisor")
        .filter(pg__in=pgs, status="pending", submitted_to_supervisor_at__lt=threshold)
        .order_by("submitted_to_supervisor_at")
    ]
    return rows, {"total": len(rows)}


def _approval_return_rate(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = list(
        LogbookEntry.objects.filter(pg__in=pgs, status__in=["approved", "returned"])
        .values(supervisor=F("supervisor__username"), department=F("rotation__department__name"))
        .annotate(
            approved=Count("id", filter=Q(status="approved")),
            returned=Count("id", filter=Q(status="returned")),
        )
        .order_by("supervisor", "department")
    )
    return rows, {"total_groups": len(rows)}


def _supervisor_productivity(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    reviewed = (
        LogbookEntry.objects.filter(pg__in=pgs, status__in=["approved", "returned", "rejected"])
        .exclude(supervisor=None)
        .values(supervisor=F("supervisor__username"))
        .annotate(
            reviewed_count=Count("id"),
            avg_turnaround_hours=Avg(F("supervisor_action_at") - F("submitted_to_supervisor_at")),
        )
    )
    rows = []
    for row in reviewed:
        duration = row["avg_turnaround_hours"]
        hours = round(duration.total_seconds() / 3600, 2) if duration else 0.0
        rows.append(
            {
                "supervisor": row["supervisor"],
                "reviewed_count": row["reviewed_count"],
                "avg_turnaround_hours": hours,
            }
        )
    return rows, {"total_supervisors": len(rows)}


def _supervisor_completion(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = []
    for sup in User.objects.filter(role="supervisor"):
        pending = LogbookEntry.objects.filter(pg__in=pgs, supervisor=sup, status="pending").count()
        reviewed = LogbookEntry.objects.filter(
            pg__in=pgs, supervisor=sup, status__in=["approved", "returned", "rejected"]
        ).count()
        rows.append(
            {
                "supervisor": sup.get_full_name(),
                "reviewed_count": reviewed,
                "pending_count": pending,
                "is_complete": pending == 0 and reviewed > 0,
            }
        )
    return rows, {"total_supervisors": len(rows)}


def _case_pipeline(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = list(
        ClinicalCase.objects.filter(pg__in=pgs, is_active=True)
        .values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )
    return rows, {"total_cases": sum(r["count"] for r in rows)}


def _cases_pending_review(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = [
        {
            "id": case.id,
            "title": case.case_title,
            "resident": case.pg.get_full_name(),
            "supervisor": case.supervisor.get_full_name() if case.supervisor else "",
            "date_encountered": case.date_encountered.isoformat(),
        }
        for case in ClinicalCase.objects.select_related("pg", "supervisor")
        .filter(pg__in=pgs, status="submitted", is_active=True)
        .order_by("-created_at")
    ]
    return rows, {"total": len(rows)}


def _case_turnaround(user, filters):
    pgs = _apply_pg_filters(_pg_scope(user), filters)
    rows = list(
        ClinicalCase.objects.filter(pg__in=pgs, reviewed_by__isnull=False, reviewed_at__isnull=False)
        .values(reviewer=F("reviewed_by__username"))
        .annotate(avg_turnaround=Avg(F("reviewed_at") - F("created_at")), reviewed_count=Count("id"))
        .order_by("reviewer")
    )
    payload = []
    for row in rows:
        duration = row["avg_turnaround"]
        payload.append(
            {
                "reviewer": row["reviewer"],
                "reviewed_count": row["reviewed_count"],
                "avg_turnaround_hours": round(duration.total_seconds() / 3600, 2) if duration else 0.0,
            }
        )
    return payload, {"total_reviewers": len(payload)}


def _bulk_import_summary(_user, _filters):
    rows = [
        {
            "created_at": op.created_at.isoformat(),
            "status": op.status,
            "success_count": op.success_count,
            "failure_count": op.failure_count,
            "actor": op.user.username,
        }
        for op in BulkOperation.objects.filter(operation=BulkOperation.OP_IMPORT).select_related("user")[:200]
    ]
    return rows, {"total_runs": len(rows)}


def _data_quality(_user, _filters):
    pg_without_supervisor = User.objects.filter(role="pg", supervisor__isnull=True).count()
    dept_without_mapping = Department.objects.exclude(
        id__in=HospitalDepartment.objects.values_list("department_id", flat=True)
    ).count()
    pending_overdue = LogbookEntry.objects.filter(
        status="pending", submitted_to_supervisor_at__lt=timezone.now() - timedelta(days=7)
    ).count()
    rows = [
        {"check": "pg_without_supervisor", "count": pg_without_supervisor},
        {"check": "department_without_hospital_mapping", "count": dept_without_mapping},
        {"check": "overdue_pending_logbook_entries", "count": pending_overdue},
    ]
    return rows, {"issues": sum(item["count"] for item in rows)}


def _activity_timeline(_user, _filters):
    rows = [
        {
            "created_at": log.created_at.isoformat(),
            "actor": log.actor.username if log.actor else "",
            "action": log.action,
            "verb": log.verb,
            "target": log.target_repr,
        }
        for log in ActivityLog.objects.select_related("actor").order_by("-created_at")[:200]
    ]
    return rows, {"total": len(rows)}


REPORT_DEFINITIONS = [
    {"key": "residents-roster", "title": "Residents roster", "runner": _residents_roster, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
    {"key": "supervisors-roster", "title": "Supervisors roster", "runner": _supervisors_roster, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
    {"key": "departments-mapping-coverage", "title": "Departments list + mapping coverage", "runner": _departments_mapping, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
    {"key": "residents-on-rotation", "title": "Residents currently on rotation", "runner": _residents_on_rotation, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "residents-no-active-rotation", "title": "Residents with no active rotation", "runner": _residents_without_active_rotation, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "upcoming-rotations", "title": "Upcoming rotations", "runner": _upcoming_rotations, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "rotation-coverage-summary", "title": "Rotation coverage summary", "runner": _rotation_coverage, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "logbook-compliance-by-resident", "title": "Logbook compliance by resident", "runner": _logbook_compliance, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "residents-logbook-up-to-date", "title": "Residents logbook up-to-date", "runner": _logbook_uptodate, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "pending-logbook-queue", "title": "Pending logbook queue", "runner": _pending_logbook_queue, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "overdue-verification", "title": "Overdue verification", "runner": _overdue_verification, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "approval-return-rate", "title": "Approval/return rate by supervisor/department", "runner": _approval_return_rate, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "supervisor-productivity-turnaround", "title": "Supervisor productivity + turnaround", "runner": _supervisor_productivity, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "supervisor-verification-completion", "title": "Supervisors 100% verification completion", "runner": _supervisor_completion, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "case-pipeline-status", "title": "Case pipeline status", "runner": _case_pipeline, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "cases-pending-review", "title": "Cases pending review", "runner": _cases_pending_review, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "case-review-turnaround", "title": "Case review turnaround", "runner": _case_turnaround, "roles": {"admin", "utrmc_admin", "utrmc_user", "supervisor"}},
    {"key": "bulk-import-run-summary", "title": "Bulk import run summary", "runner": _bulk_import_summary, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
    {"key": "data-quality-checks", "title": "Data quality checks", "runner": _data_quality, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
    {"key": "activity-timeline", "title": "Activity timeline", "runner": _activity_timeline, "roles": {"admin", "utrmc_admin", "utrmc_user"}},
]


def report_catalog_for(user) -> list:
    return [
        {"key": item["key"], "title": item["title"]}
        for item in REPORT_DEFINITIONS
        if user.role in item["roles"] or user.is_superuser
    ]


def run_report_for(user, key: str, filters: dict) -> Tuple[Rows, Summary]:
    definition = next((item for item in REPORT_DEFINITIONS if item["key"] == key), None)
    if definition is None:
        raise KeyError("Unknown report key.")
    if not (user.is_superuser or user.role in definition["roles"]):
        raise PermissionError("You do not have permission to run this report.")
    return definition["runner"](user, filters)


def export_rows(rows: Rows, key: str, export_format: str) -> Tuple[bytes, str, str]:
    if export_format == "csv":
        output = io.StringIO()
        headers = list(rows[0].keys()) if rows else []
        writer = csv.DictWriter(output, fieldnames=headers)
        if headers:
            writer.writeheader()
            writer.writerows(rows)
        return output.getvalue().encode("utf-8"), "text/csv", f"{key}.csv"

    workbook = Workbook()
    sheet = workbook.active
    headers = list(rows[0].keys()) if rows else []
    for idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
    stream = io.BytesIO()
    workbook.save(stream)
    return (
        stream.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{key}.xlsx",
    )
