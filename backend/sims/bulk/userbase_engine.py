"""Active-surface userbase bulk import/export helpers."""

from __future__ import annotations

import csv
import io
import re
import secrets
import string
from datetime import date, datetime
from typing import Iterable, List, Tuple

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from sims.academics.models import Department
from sims.rotations.models import Hospital, HospitalDepartment
from sims.users.models import (
    DepartmentMembership,
    HODAssignment,
    HospitalAssignment,
    ResidentProfile,
    SPECIALTY_CHOICES,
    StaffProfile,
    SupervisorResidentLink,
    YEAR_CHOICES,
)

User = get_user_model()

SUPPORTED_IMPORT_ENTITIES = {
    "hospitals",
    "departments",
    "matrix",
    "faculty-supervisors",
    "residents",
    "supervision-links",
    "hod-assignments",
}

SUPPORTED_EXPORT_RESOURCES = SUPPORTED_IMPORT_ENTITIES

TEMPLATE_ROWS = {
    "hospitals": [
        {
            "hospital_code": "AH",
            "hospital_name": "Allied Hospital",
            "address": "Faisalabad",
            "phone": "0410000000",
            "email": "allied@example.com",
            "active": "true",
        }
    ],
    "departments": [
        {
            "department_code": "MED",
            "department_name": "Internal Medicine",
            "description": "Core medicine department",
            "active": "true",
        }
    ],
    "matrix": [
        {
            "hospital_code": "AH",
            "department_code": "MED",
            "active": "true",
        }
    ],
    "faculty-supervisors": [
        {
            "email": "supervisor@example.com",
            "full_name": "Dr. Jane Supervisor",
            "phone_number": "03001234567",
            "role": "supervisor",
            "specialty": "medicine",
            "department_code": "MED",
            "hospital_code": "AH",
            "designation": "Consultant",
            "registration_number": "PMC-12345",
            "username": "jane.supervisor",
            "password": "",
            "active": "true",
            "start_date": "2026-01-01",
        }
    ],
    "residents": [
        {
            "email": "resident@example.com",
            "full_name": "Dr. Ali Resident",
            "phone_number": "03009876543",
            "role": "resident",
            "specialty": "medicine",
            "year": "1",
            "pgr_id": "PGR001",
            "training_start": "2026-01-01",
            "training_end": "",
            "training_level": "Y1",
            "department_code": "MED",
            "hospital_code": "AH",
            "supervisor_email": "supervisor@example.com",
            "username": "ali.resident",
            "password": "",
            "active": "true",
        }
    ],
    "supervision-links": [
        {
            "supervisor_email": "supervisor@example.com",
            "resident_email": "resident@example.com",
            "department_code": "MED",
            "start_date": "2026-01-01",
            "end_date": "",
            "active": "true",
        }
    ],
    "hod-assignments": [
        {
            "department_code": "MED",
            "hod_email": "supervisor@example.com",
            "start_date": "2026-01-01",
            "end_date": "",
            "active": "true",
        }
    ],
}


def parse_tabular_rows(uploaded_file) -> List[dict]:
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    stream = io.BytesIO(content) if isinstance(content, bytes) else io.StringIO(content)
    stream.seek(0)

    if name.endswith(".csv"):
        text_stream = io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
        reader = csv.DictReader(text_stream)
        headers = [_normalize_header(header or "") for header in (reader.fieldnames or [])]
        reader.fieldnames = headers
        return [
            {
                **{
                    _normalize_header(key or ""): (value.strip() if isinstance(value, str) else str(value or "").strip())
                    for key, value in row.items()
                    if key
                },
                "_row_number": index,
            }
            for index, row in enumerate(reader, start=2)
            if any((value or "").strip() for value in row.values())
        ]

    if name.endswith((".xlsx", ".xls")):
        workbook = load_workbook(stream)
        sheet = workbook.active
        headers = [
            _normalize_header(str(cell.value)) if cell.value is not None else f"col_{index}"
            for index, cell in enumerate(next(sheet.iter_rows(max_row=1)))
        ]
        rows: List[dict] = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            payload = {
                headers[idx]: (str(value).strip() if value is not None else "")
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            payload["_row_number"] = index
            rows.append(payload)
        return rows

    raise ValidationError("Unsupported file format. Upload CSV or Excel.")


def template_rows_for(resource: str) -> List[dict]:
    if resource not in TEMPLATE_ROWS:
        raise ValidationError(f"Unsupported template resource '{resource}'.")
    return TEMPLATE_ROWS[resource]


def export_rows_for(resource: str) -> List[dict]:
    if resource == "hospitals":
        return [
            {
                "hospital_code": hospital.code or "",
                "hospital_name": hospital.name,
                "address": hospital.address or "",
                "phone": hospital.phone or "",
                "email": hospital.email or "",
                "active": _bool_str(hospital.is_active),
            }
            for hospital in Hospital.objects.all().order_by("name")
        ]

    if resource == "departments":
        return [
            {
                "department_code": department.code,
                "department_name": department.name,
                "description": department.description or "",
                "active": _bool_str(department.active),
            }
            for department in Department.objects.all().order_by("name")
        ]

    if resource == "matrix":
        return [
            {
                "hospital_code": item.hospital.code or "",
                "department_code": item.department.code,
                "active": _bool_str(item.is_active),
            }
            for item in HospitalDepartment.objects.select_related("hospital", "department").order_by(
                "hospital__name",
                "department__name",
            )
        ]

    if resource == "faculty-supervisors":
        rows: List[dict] = []
        staff_users = User.objects.filter(role__in=["faculty", "supervisor"]).order_by("last_name", "first_name")
        for user in staff_users:
            membership = _primary_membership_for(user, {"faculty", "supervisor"})
            assignment = _active_assignment_for(user, HospitalAssignment.ASSIGNMENT_FACULTY_SITE)
            profile = getattr(user, "staff_profile", None)
            rows.append(
                {
                    "email": user.email,
                    "full_name": user.get_full_name(),
                    "phone_number": user.phone_number or "",
                    "role": user.role,
                    "specialty": user.specialty or "",
                    "department_code": membership.department.code if membership else "",
                    "hospital_code": assignment.hospital_department.hospital.code if assignment else "",
                    "designation": profile.designation if profile else "",
                    "registration_number": user.registration_number or "",
                    "username": user.username,
                    "password": "",
                    "active": _bool_str(user.is_active),
                    "start_date": membership.start_date.isoformat() if membership else "",
                }
            )
        return rows

    if resource == "residents":
        rows: List[dict] = []
        resident_users = User.objects.filter(role__in=["resident", "pg"]).order_by("last_name", "first_name")
        for user in resident_users:
            membership = _primary_membership_for(user, {"resident"})
            assignment = _active_assignment_for(user, HospitalAssignment.ASSIGNMENT_PRIMARY_TRAINING)
            profile = getattr(user, "resident_profile", None)
            rows.append(
                {
                    "email": user.email,
                    "full_name": user.get_full_name(),
                    "phone_number": user.phone_number or "",
                    "role": user.role,
                    "specialty": user.specialty or "",
                    "year": user.year or "",
                    "pgr_id": profile.pgr_id if profile else "",
                    "training_start": profile.training_start.isoformat() if profile else "",
                    "training_end": profile.training_end.isoformat() if profile and profile.training_end else "",
                    "training_level": profile.training_level if profile else "",
                    "department_code": membership.department.code if membership else "",
                    "hospital_code": assignment.hospital_department.hospital.code if assignment else "",
                    "supervisor_email": user.supervisor.email if user.supervisor else "",
                    "username": user.username,
                    "password": "",
                    "active": _bool_str(user.is_active),
                }
            )
        return rows

    if resource == "supervision-links":
        return [
            {
                "supervisor_email": link.supervisor_user.email,
                "resident_email": link.resident_user.email,
                "department_code": link.department.code if link.department else "",
                "start_date": link.start_date.isoformat(),
                "end_date": link.end_date.isoformat() if link.end_date else "",
                "active": _bool_str(link.active),
            }
            for link in SupervisorResidentLink.objects.select_related(
                "supervisor_user",
                "resident_user",
                "department",
            ).order_by("resident_user__last_name", "resident_user__first_name")
        ]

    if resource == "hod-assignments":
        return [
            {
                "department_code": assignment.department.code,
                "hod_email": assignment.hod_user.email,
                "start_date": assignment.start_date.isoformat(),
                "end_date": assignment.end_date.isoformat() if assignment.end_date else "",
                "active": _bool_str(assignment.active),
            }
            for assignment in HODAssignment.objects.select_related("department", "hod_user").order_by(
                "department__name",
                "-start_date",
            )
        ]

    raise ValidationError(f"Unsupported export resource '{resource}'.")


def import_entity(actor: User, entity: str, uploaded_file, *, dry_run: bool, allow_partial: bool) -> dict:
    if entity not in SUPPORTED_IMPORT_ENTITIES:
        raise ValidationError(f"Unsupported import entity '{entity}'.")
    rows = parse_tabular_rows(uploaded_file)
    if not rows:
        raise ValidationError("No data rows found in file.")

    handlers = {
        "hospitals": _import_hospitals,
        "departments": _import_departments,
        "matrix": _import_matrix,
        "faculty-supervisors": _import_faculty_supervisors,
        "residents": _import_residents,
        "supervision-links": _import_supervision_links,
        "hod-assignments": _import_hod_assignments,
    }
    return handlers[entity](actor, rows, dry_run=dry_run, allow_partial=allow_partial)


def _import_hospitals(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        code = (row.get("hospital_code") or row.get("code") or "").strip().upper()
        name = (row.get("hospital_name") or row.get("name") or "").strip()
        if not code or not name:
            failures.append({"row": row_number, "error": "hospital_code and hospital_name are required"})
            if not allow_partial:
                break
            continue
        try:
            if not dry_run:
                Hospital.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "address": (row.get("address") or "").strip(),
                        "phone": (row.get("phone") or "").strip(),
                        "email": (row.get("email") or "").strip(),
                        "is_active": _parse_bool(row.get("active"), default=True),
                    },
                )
            successes.append({"row": row_number, "hospital_code": code, "hospital_name": name})
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _import_departments(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        code = (row.get("department_code") or row.get("code") or "").strip().upper()
        name = (row.get("department_name") or row.get("name") or "").strip()
        if not code or not name:
            failures.append({"row": row_number, "error": "department_code and department_name are required"})
            if not allow_partial:
                break
            continue
        try:
            if dry_run:
                candidate = Department(code=code, name=name)
                candidate.description = (row.get("description") or "").strip()
                candidate.active = _parse_bool(row.get("active"), default=True)
                candidate.full_clean()
            else:
                Department.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "description": (row.get("description") or "").strip(),
                        "active": _parse_bool(row.get("active"), default=True),
                    },
                )
            successes.append({"row": row_number, "department_code": code, "department_name": name})
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _import_matrix(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        hospital_code = (row.get("hospital_code") or "").strip().upper()
        department_code = (row.get("department_code") or "").strip().upper()
        if not hospital_code or not department_code:
            failures.append({"row": row_number, "error": "hospital_code and department_code are required"})
            if not allow_partial:
                break
            continue
        hospital = Hospital.objects.filter(code=hospital_code).first()
        department = Department.objects.filter(code=department_code).first()
        if not hospital or not department:
            missing = []
            if not hospital:
                missing.append(f"hospital '{hospital_code}'")
            if not department:
                missing.append(f"department '{department_code}'")
            failures.append({"row": row_number, "error": f"Missing prerequisite: {', '.join(missing)}"})
            if not allow_partial:
                break
            continue
        if not dry_run:
            HospitalDepartment.objects.update_or_create(
                hospital=hospital,
                department=department,
                defaults={"is_active": _parse_bool(row.get("active"), default=True)},
            )
        successes.append({"row": row_number, "hospital_code": hospital_code, "department_code": department_code})
    return {"successes": successes, "failures": failures}


def _import_faculty_supervisors(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        try:
            role = _normalize_staff_role(row.get("role"))
            specialty = _normalize_specialty(row.get("specialty"))
            first_name, last_name = _parse_person_name(row)
            email = _require_text(row, "email")
            username = (row.get("username") or _generate_username(first_name, last_name)).strip()
            active = _parse_bool(row.get("active"), default=True)
            start_date = _parse_date_value(row.get("start_date")) or date.today()
            department = _department_or_none(row.get("department_code"))
            hospital_department = _resolve_hospital_department(
                hospital_code=row.get("hospital_code"),
                department=department,
            )

            with transaction.atomic():
                user = _build_or_load_user(email=email, username=username)
                _assert_username_available(user, username)
                if dry_run:
                    candidate = user or User(username=username)
                    candidate.email = email
                    candidate.first_name = first_name
                    candidate.last_name = last_name
                    candidate.role = role
                    candidate.specialty = specialty
                    candidate.phone_number = (row.get("phone_number") or row.get("phone") or "").strip() or None
                    candidate.registration_number = (row.get("registration_number") or "").strip() or None
                    candidate.is_active = active
                    candidate.full_clean()
                else:
                    user, generated_password = _upsert_staff_user(
                        actor=actor,
                        existing=user,
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        role=role,
                        specialty=specialty,
                        phone_number=(row.get("phone_number") or row.get("phone") or "").strip(),
                        registration_number=(row.get("registration_number") or "").strip(),
                        designation=(row.get("designation") or "").strip(),
                        password=(row.get("password") or "").strip(),
                        active=active,
                    )
                    if department:
                        _sync_primary_membership(
                            actor=actor,
                            user=user,
                            department=department,
                            member_type=DepartmentMembership.MEMBER_SUPERVISOR if role == "supervisor" else DepartmentMembership.MEMBER_FACULTY,
                            start_date=start_date,
                            active=active,
                        )
                    if hospital_department:
                        _sync_assignment(
                            actor=actor,
                            user=user,
                            hospital_department=hospital_department,
                            assignment_type=HospitalAssignment.ASSIGNMENT_FACULTY_SITE,
                            start_date=start_date,
                            active=active,
                        )
                success = {
                    "row": row_number,
                    "email": email,
                    "role": role,
                    "department_code": department.code if department else "",
                    "hospital_code": hospital_department.hospital.code if hospital_department else "",
                }
                if not dry_run and generated_password:
                    success["temporary_password"] = generated_password
                successes.append(success)
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _import_residents(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        try:
            role = _normalize_resident_role(row.get("role"))
            specialty = _normalize_specialty(row.get("specialty"))
            year = _normalize_year(row.get("year"))
            training_start = _parse_required_date(row, "training_start")
            training_end = _parse_date_value(row.get("training_end"))
            first_name, last_name = _parse_person_name(row)
            email = _require_text(row, "email")
            username = (row.get("username") or _generate_username(first_name, last_name)).strip()
            active = _parse_bool(row.get("active"), default=True)
            department = _department_or_none(row.get("department_code"))
            hospital_department = _resolve_hospital_department(
                hospital_code=row.get("hospital_code"),
                department=department,
            )
            supervisor = _user_by_email_or_none(row.get("supervisor_email"), {"supervisor", "faculty"})

            with transaction.atomic():
                user = _build_or_load_user(email=email, username=username)
                _assert_username_available(user, username)
                if dry_run:
                    candidate = user or User(username=username)
                    candidate.email = email
                    candidate.first_name = first_name
                    candidate.last_name = last_name
                    candidate.role = role
                    candidate.specialty = specialty
                    candidate.year = year
                    candidate.phone_number = (row.get("phone_number") or row.get("phone") or "").strip() or None
                    candidate.registration_number = (row.get("registration_number") or row.get("pgr_id") or "").strip() or None
                    candidate.is_active = active
                    candidate.home_department = department
                    candidate.home_hospital = hospital_department.hospital if hospital_department else None
                    candidate.supervisor = supervisor
                    candidate.full_clean()
                else:
                    user, generated_password = _upsert_resident_user(
                        actor=actor,
                        existing=user,
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        role=role,
                        specialty=specialty,
                        year=year,
                        phone_number=(row.get("phone_number") or row.get("phone") or "").strip(),
                        registration_number=(row.get("registration_number") or "").strip(),
                        pgr_id=(row.get("pgr_id") or "").strip(),
                        training_start=training_start,
                        training_end=training_end,
                        training_level=(row.get("training_level") or "").strip(),
                        department=department,
                        hospital_department=hospital_department,
                        supervisor=supervisor,
                        password=(row.get("password") or "").strip(),
                        active=active,
                    )
                    if department:
                        _sync_primary_membership(
                            actor=actor,
                            user=user,
                            department=department,
                            member_type=DepartmentMembership.MEMBER_RESIDENT,
                            start_date=training_start,
                            active=active,
                        )
                    if hospital_department:
                        _sync_assignment(
                            actor=actor,
                            user=user,
                            hospital_department=hospital_department,
                            assignment_type=HospitalAssignment.ASSIGNMENT_PRIMARY_TRAINING,
                            start_date=training_start,
                            active=active,
                        )
                success = {
                    "row": row_number,
                    "email": email,
                    "role": role,
                    "department_code": department.code if department else "",
                    "hospital_code": hospital_department.hospital.code if hospital_department else "",
                }
                if not dry_run and generated_password:
                    success["temporary_password"] = generated_password
                successes.append(success)
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _import_supervision_links(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        try:
            supervisor = _required_user_by_email(row, "supervisor_email", {"supervisor", "faculty"})
            resident = _required_user_by_email(row, "resident_email", {"resident", "pg"})
            department = _department_or_none(row.get("department_code"))
            start_date = _parse_required_date(row, "start_date")
            end_date = _parse_date_value(row.get("end_date"))
            active = _parse_bool(row.get("active"), default=True)
            if not dry_run:
                SupervisorResidentLink.objects.update_or_create(
                    supervisor_user=supervisor,
                    resident_user=resident,
                    department=department,
                    defaults={
                        "start_date": start_date,
                        "end_date": end_date,
                        "active": active,
                        "created_by": actor,
                        "updated_by": actor,
                    },
                )
            successes.append(
                {
                    "row": row_number,
                    "supervisor_email": supervisor.email,
                    "resident_email": resident.email,
                    "department_code": department.code if department else "",
                }
            )
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _import_hod_assignments(actor: User, rows: List[dict], *, dry_run: bool, allow_partial: bool) -> dict:
    successes: List[dict] = []
    failures: List[dict] = []
    for row in rows:
        row_number = row["_row_number"]
        try:
            department = _required_department(row)
            hod_user = _required_user_by_email(row, "hod_email", {"supervisor", "faculty"})
            start_date = _parse_required_date(row, "start_date")
            end_date = _parse_date_value(row.get("end_date"))
            active = _parse_bool(row.get("active"), default=True)
            if not dry_run:
                current = HODAssignment.objects.filter(department=department, active=True).first()
                if current:
                    current.hod_user = hod_user
                    current.start_date = start_date
                    current.end_date = end_date
                    current.active = active
                    current.updated_by = actor
                    current.full_clean()
                    current.save()
                else:
                    HODAssignment.objects.create(
                        department=department,
                        hod_user=hod_user,
                        start_date=start_date,
                        end_date=end_date,
                        active=active,
                        created_by=actor,
                        updated_by=actor,
                    )
            successes.append({"row": row_number, "department_code": department.code, "hod_email": hod_user.email})
        except ValidationError as exc:
            failures.append({"row": row_number, "error": _error_text(exc)})
            if not allow_partial:
                break
    return {"successes": successes, "failures": failures}


def _upsert_staff_user(
    *,
    actor: User,
    existing: User | None,
    username: str,
    email: str,
    first_name: str,
    last_name: str,
    role: str,
    specialty: str | None,
    phone_number: str,
    registration_number: str,
    designation: str,
    password: str,
    active: bool,
) -> Tuple[User, str | None]:
    generated_password = None
    user = existing or User(username=username)
    user.email = email
    user.username = username
    user.first_name = first_name
    user.last_name = last_name
    user.role = role
    user.specialty = specialty
    user.phone_number = phone_number or None
    user.registration_number = registration_number or None
    user.is_active = active
    if existing:
        user.modified_by = actor
    else:
        user.created_by = actor
        user.modified_by = actor
    if password:
        user.set_password(password)
    elif not existing:
        generated_password = _generate_secure_password()
        user.set_password(generated_password)
    user.full_clean()
    user.save()
    StaffProfile.objects.update_or_create(
        user=user,
        defaults={
            "designation": designation,
            "phone": phone_number,
            "active": active,
        },
    )
    return user, generated_password


def _upsert_resident_user(
    *,
    actor: User,
    existing: User | None,
    username: str,
    email: str,
    first_name: str,
    last_name: str,
    role: str,
    specialty: str,
    year: str,
    phone_number: str,
    registration_number: str,
    pgr_id: str,
    training_start: date,
    training_end: date | None,
    training_level: str,
    department: Department | None,
    hospital_department: HospitalDepartment | None,
    supervisor: User | None,
    password: str,
    active: bool,
) -> Tuple[User, str | None]:
    generated_password = None
    user = existing or User(username=username)
    user.email = email
    user.username = username
    user.first_name = first_name
    user.last_name = last_name
    user.role = role
    user.specialty = specialty
    user.year = year
    user.phone_number = phone_number or None
    user.registration_number = registration_number or None
    user.home_department = department
    user.home_hospital = hospital_department.hospital if hospital_department else None
    user.supervisor = supervisor
    user.is_active = active
    if existing:
        user.modified_by = actor
    else:
        user.created_by = actor
        user.modified_by = actor
    if password:
        user.set_password(password)
    elif not existing:
        generated_password = _generate_secure_password()
        user.set_password(generated_password)
    user.full_clean()
    user.save()
    ResidentProfile.objects.update_or_create(
        user=user,
        defaults={
            "pgr_id": pgr_id,
            "training_start": training_start,
            "training_end": training_end,
            "training_level": training_level,
            "active": active,
        },
    )
    return user, generated_password


def _sync_primary_membership(
    *,
    actor: User,
    user: User,
    department: Department,
    member_type: str,
    start_date: date,
    active: bool,
) -> None:
    DepartmentMembership.objects.filter(
        user=user,
        is_primary=True,
        active=True,
    ).exclude(department=department).update(
        active=False,
        end_date=start_date,
        updated_by=actor,
    )
    DepartmentMembership.objects.update_or_create(
        user=user,
        department=department,
        defaults={
            "member_type": member_type,
            "is_primary": True,
            "active": active,
            "start_date": start_date,
            "end_date": None,
            "created_by": actor,
            "updated_by": actor,
        },
    )


def _sync_assignment(
    *,
    actor: User,
    user: User,
    hospital_department: HospitalDepartment,
    assignment_type: str,
    start_date: date,
    active: bool,
) -> None:
    HospitalAssignment.objects.filter(
        user=user,
        assignment_type=assignment_type,
        active=True,
    ).exclude(hospital_department=hospital_department).update(
        active=False,
        end_date=start_date,
        updated_by=actor,
    )
    existing = HospitalAssignment.objects.filter(
        user=user,
        hospital_department=hospital_department,
        assignment_type=assignment_type,
    ).order_by("-active", "-start_date").first()
    if existing:
        existing.start_date = start_date
        existing.end_date = None
        existing.active = active
        existing.updated_by = actor
        existing.full_clean()
        existing.save()
        return
    HospitalAssignment.objects.create(
        user=user,
        hospital_department=hospital_department,
        assignment_type=assignment_type,
        start_date=start_date,
        active=active,
        created_by=actor,
        updated_by=actor,
    )


def _primary_membership_for(user: User, member_types: set[str]) -> DepartmentMembership | None:
    return (
        user.department_memberships.filter(active=True, is_primary=True, member_type__in=member_types)
        .select_related("department")
        .order_by("-start_date")
        .first()
    )


def _active_assignment_for(user: User, assignment_type: str) -> HospitalAssignment | None:
    return (
        user.hospital_assignments.filter(active=True, assignment_type=assignment_type)
        .select_related("hospital_department", "hospital_department__hospital")
        .order_by("-start_date")
        .first()
    )


def _build_or_load_user(*, email: str, username: str) -> User | None:
    if username:
        user = User.objects.filter(username=username).first()
        if user:
            return user
    if email:
        return User.objects.filter(email__iexact=email).first()
    return None


def _assert_username_available(user: User | None, username: str) -> None:
    conflict = User.objects.filter(username=username)
    if user:
        conflict = conflict.exclude(pk=user.pk)
    if conflict.exists():
        raise ValidationError({"username": f"Username '{username}' is already in use."})


def _required_department(row: dict) -> Department:
    return _resolve_department(_require_text(row, "department_code"))


def _department_or_none(raw_code: str | None) -> Department | None:
    code = (raw_code or "").strip().upper()
    if not code:
        return None
    return _resolve_department(code)


def _resolve_department(code: str) -> Department:
    department = Department.objects.filter(code=code).first()
    if not department:
        raise ValidationError({"department_code": f"Department '{code}' not found."})
    return department


def _resolve_hospital_department(*, hospital_code: str | None, department: Department | None) -> HospitalDepartment | None:
    code = (hospital_code or "").strip().upper()
    if not code:
        return None
    if not department:
        raise ValidationError({"hospital_code": "hospital_code requires department_code so the active matrix site can be resolved."})
    hospital = Hospital.objects.filter(code=code).first()
    if not hospital:
        raise ValidationError({"hospital_code": f"Hospital '{code}' not found."})
    hospital_department = HospitalDepartment.objects.filter(hospital=hospital, department=department).first()
    if not hospital_department:
        raise ValidationError({"hospital_code": f"Hospital '{code}' is not linked to department '{department.code}' in the matrix."})
    return hospital_department


def _required_user_by_email(row: dict, field: str, allowed_roles: set[str]) -> User:
    return _resolve_user_by_email(_require_text(row, field), field, allowed_roles)


def _user_by_email_or_none(raw_email: str | None, allowed_roles: set[str]) -> User | None:
    email = (raw_email or "").strip().lower()
    if not email:
        return None
    return _resolve_user_by_email(email, "email", allowed_roles)


def _resolve_user_by_email(email: str, field: str, allowed_roles: set[str]) -> User:
    user = User.objects.filter(email__iexact=email).first()
    if not user:
        raise ValidationError({field: f"User '{email}' not found."})
    if user.role not in allowed_roles:
        raise ValidationError({field: f"User '{email}' must have one of: {', '.join(sorted(allowed_roles))}."})
    return user


def _parse_person_name(row: dict) -> Tuple[str, str]:
    full_name = (row.get("full_name") or "").strip()
    if full_name:
        return _split_name(full_name)
    first_name = (row.get("first_name") or "").strip()
    last_name = (row.get("last_name") or "").strip()
    if not first_name:
        raise ValidationError({"full_name": "Provide full_name or first_name."})
    return first_name, last_name


def _split_name(full_name: str) -> Tuple[str, str]:
    cleaned = re.sub(r"^(dr\.?|prof\.?|mr\.?|mrs\.?|ms\.?)\s+", "", full_name.strip(), flags=re.IGNORECASE)
    parts = [part for part in cleaned.split() if part]
    if not parts:
        raise ValidationError({"full_name": "Invalid full_name."})
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def _normalize_staff_role(raw_role: str | None) -> str:
    role = (raw_role or "").strip().lower()
    if role not in {"faculty", "supervisor"}:
        raise ValidationError({"role": "Role must be faculty or supervisor."})
    return role


def _normalize_resident_role(raw_role: str | None) -> str:
    role = (raw_role or "resident").strip().lower()
    if role not in {"resident", "pg"}:
        raise ValidationError({"role": "Role must be resident or pg."})
    return role


def _normalize_year(raw_year: str | None) -> str:
    year = (raw_year or "").strip()
    valid = {choice[0] for choice in YEAR_CHOICES}
    if year not in valid:
        raise ValidationError({"year": f"Year must be one of: {', '.join(sorted(valid))}."})
    return year


def _normalize_specialty(raw_specialty: str | None) -> str | None:
    specialty = (raw_specialty or "").strip()
    if not specialty:
        return None
    normalized = specialty.lower().replace(" ", "_")
    valid = {choice[0] for choice in SPECIALTY_CHOICES}
    if normalized in valid:
        return normalized
    for code, label in SPECIALTY_CHOICES:
        if specialty.lower() == label.lower():
            return code
    raise ValidationError({"specialty": f"Unknown specialty '{specialty}'."})


def _parse_required_date(row: dict, field: str) -> date:
    raw = _require_text(row, field)
    parsed = _parse_date_value(raw)
    if not parsed:
        raise ValidationError({field: f"{field} is required."})
    return parsed


def _parse_date_value(raw_value: str | None) -> date | None:
    raw = (raw_value or "").strip()
    if not raw:
        return None
    formats = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")
    for fmt in formats:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValidationError({"date": f"Invalid date '{raw}'. Use YYYY-MM-DD."})


def _parse_bool(raw_value: str | None, *, default: bool) -> bool:
    raw = (raw_value or "").strip().lower()
    if not raw:
        return default
    if raw in {"true", "1", "yes", "y"}:
        return True
    if raw in {"false", "0", "no", "n"}:
        return False
    raise ValidationError({"active": f"Invalid boolean '{raw_value}'."})


def _bool_str(value: bool) -> str:
    return "true" if value else "false"


def _generate_username(first_name: str, last_name: str) -> str:
    base = ".".join(part for part in [_slug(first_name), _slug(last_name)] if part) or "user"
    candidate = base
    counter = 1
    while User.objects.filter(username=candidate).exists():
        counter += 1
        candidate = f"{base}{counter}"
    return candidate


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _generate_secure_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _require_text(row: dict, field: str) -> str:
    value = (row.get(field) or "").strip()
    if not value:
        raise ValidationError({field: f"{field} is required."})
    return value


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_")


def _error_text(exc: ValidationError) -> str:
    if hasattr(exc, "message_dict"):
        parts: List[str] = []
        for key, messages in exc.message_dict.items():
            if isinstance(messages, Iterable) and not isinstance(messages, str):
                parts.append(f"{key}: {', '.join(str(message) for message in messages)}")
            else:
                parts.append(f"{key}: {messages}")
        return "; ".join(parts)
    if hasattr(exc, "messages"):
        return "; ".join(str(message) for message in exc.messages)
    return str(exc)
