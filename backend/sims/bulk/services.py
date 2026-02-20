"""Service layer for bulk operations."""

from __future__ import annotations

import csv
import io
import re
import secrets
import string
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Iterable, Iterator, List, Optional, Sequence, Tuple

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from sims.bulk.models import BulkOperation
from sims.logbook.models import LogbookEntry
from sims.users.models import SPECIALTY_CHOICES, YEAR_CHOICES, User

try:
    from sims.rotations.models import Department, Hospital
except ImportError:
    Department = None
    Hospital = None


@dataclass
class BulkResult:
    successes: List[dict]
    failures: List[dict]


class BulkProcessingError(Exception):
    """Raised when a bulk operation fails in transactional mode."""


class BulkService:
    def __init__(self, actor: User, chunk_size: int = 50):
        self.actor = actor
        self.chunk_size = chunk_size
        self._validate_permissions()

    def _validate_permissions(self) -> None:
        if not (
            self.actor.is_superuser or getattr(self.actor, "role", None) in {"admin", "supervisor"}
        ):
            raise PermissionDenied("Bulk operations are restricted to supervisors and admins.")

    # ------------------------------------------------------------------
    # Review and assignment operations

    def review_entries(self, entry_ids: Sequence[int], status: str) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_REVIEW)
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.status = status
                        entry.supervisor_action_at = timezone.now()
                        entry.save(update_fields=["status", "supervisor_action_at"])
                        successes.append({"id": entry.pk, "status": status})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    def assign_supervisor(self, entry_ids: Sequence[int], supervisor: User) -> BulkOperation:
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_ASSIGNMENT
        )
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.supervisor = supervisor
                        entry.save(update_fields=["supervisor"])
                        successes.append({"id": entry.pk, "supervisor": supervisor.pk})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    # ------------------------------------------------------------------
    # Bulk import

    def import_logbook_entries(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
    ) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_IMPORT)
        rows = list(_parse_rows(uploaded_file))
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False

        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            try:
                pg = User.objects.get(username=row["pg_username"], role="pg")
            except User.DoesNotExist:
                failures.append({"row": row, "error": "invalid-pg"})
                errors_triggered = True
                return
            try:
                entry_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
            except (KeyError, ValueError):
                failures.append({"row": row, "error": "invalid-date"})
                errors_triggered = True
                return
            status = row.get("status", "draft")
            payload = {
                "pg": pg,
                "case_title": row.get("case_title") or "Untitled",
                "date": entry_date,
                "status": status,
                "location_of_activity": row.get("location") or "Not specified",
                "patient_history_summary": row.get("patient_history") or "Pending summary",
                "management_action": row.get("management_action") or "Pending action",
                "topic_subtopic": row.get("topic_subtopic") or "General",
            }
            try:
                if dry_run:
                    entry = LogbookEntry(**payload)
                    entry.full_clean()
                else:
                    entry = LogbookEntry.objects.create(**payload, created_by=self.actor)
                successes.append(
                    {
                        "pg": pg.username,
                        "case_title": payload["case_title"],
                        "status": status,
                    }
                )
            except ValidationError as exc:
                failures.append({"row": row, "error": exc.message_dict})
                errors_triggered = True

        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation

        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    def import_supervisors(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
        generate_passwords: bool = True,
    ) -> BulkOperation:
        """
        Import supervisor/faculty data from CSV or Excel file.
        
        Expected columns (case-insensitive):
        - Name (or First Name, Last Name)
        - Email (optional, will be generated)
        - Specialty (required)
        - Department (optional)
        - Phone (optional)
        - Registration Number (optional)
        - Username (optional, will be generated)
        
        Creates accounts with role 'supervisor' and generates secure passwords.
        """
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_IMPORT
        )
        
        # Required columns (flexible matching)
        required_cols = {"name", "specialty"}  # Flexible - can be "first name" + "last name"
        
        try:
            rows = list(_parse_csv_rows(uploaded_file, required_columns=None))
        except ValidationError as e:
            operation.mark_failed({"error": str(e)})
            return operation
        
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False
        
        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            
            row_num = row.get("_row_number", "unknown")
            
            # Extract data with flexible column matching
            name = (
                row.get("name") or 
                f"{row.get('first_name', '').strip()} {row.get('last_name', '').strip()}".strip()
            )
            email = row.get("email", "").strip()
            specialty = row.get("specialty", "").strip()
            department_name = row.get("department", "").strip()
            phone = row.get("phone", "").strip() or row.get("phone_number", "").strip()
            registration_number = row.get("registration_number", "").strip() or row.get("reg_no", "").strip()
            username = row.get("username", "").strip()
            
            # Validate required fields
            if not name:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Name' (or 'First Name' + 'Last Name')",
                    "data": row
                })
                errors_triggered = True
                return
            
            if not specialty:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Specialty'",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Validate specialty
            valid_specialties = [choice[0] for choice in SPECIALTY_CHOICES]
            specialty_lower = specialty.lower().replace(" ", "_")
            # Try to match specialty
            matched_specialty = None
            for spec_code, spec_name in SPECIALTY_CHOICES:
                if (specialty_lower == spec_code.lower() or 
                    specialty.lower() == spec_name.lower() or
                    specialty_lower in spec_name.lower() or
                    spec_name.lower() in specialty_lower):
                    matched_specialty = spec_code
                    break
            
            if not matched_specialty:
                failures.append({
                    "row": row_num,
                    "error": f"Invalid specialty '{specialty}'. Valid options: {', '.join([s[1] for s in SPECIALTY_CHOICES[:5]])}...",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse name
            try:
                first_name, last_name = _parse_name(name)
                if not first_name and not last_name:
                    failures.append({
                        "row": row_num,
                        "error": "Invalid name format",
                        "data": row
                    })
                    errors_triggered = True
                    return
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error parsing name: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Generate username if not provided
            if not username:
                try:
                    username = _generate_username(first_name, last_name)
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "error": f"Error generating username: {str(e)}",
                        "data": row
                    })
                    errors_triggered = True
                    return
            
            # Generate email if not provided
            if not email:
                email = f"{username}.supervisor@pmc.edu.pk"
            
            # Get or create department if provided
            department = None
            if department_name and Department:
                try:
                    # Try to find department by name (case-insensitive)
                    department = Department.objects.filter(
                        name__iexact=department_name,
                        is_active=True
                    ).first()
                    
                    if not department:
                        # Optionally create department if hospital exists
                        # For now, we'll just warn
                        if allow_partial:
                            pass  # Continue without department
                        else:
                            failures.append({
                                "row": row_num,
                                "warning": f"Department '{department_name}' not found. Continuing without department.",
                                "data": row
                            })
                except Exception as e:
                    if not allow_partial:
                        failures.append({
                            "row": row_num,
                            "error": f"Error with department '{department_name}': {str(e)}",
                            "data": row
                        })
                        errors_triggered = True
                        return
            
            # Prepare user data
            user_data = {
                "username": username,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": "supervisor",
                "specialty": matched_specialty,
                "phone_number": phone if phone else None,
                "registration_number": registration_number if registration_number else None,
                "is_active": True,
                "created_by": self.actor,
            }
            
            # Generate password
            password = None
            if generate_passwords:
                password = _generate_password_from_username(username)
            else:
                password = _generate_secure_password()
            
            # Create or validate user
            try:
                if dry_run:
                    # Validate without creating
                    user = User(**user_data)
                    user.full_clean()
                else:
                    # Check if user already exists
                    existing_user = User.objects.filter(username=username).first()
                    if existing_user:
                        # Update existing user
                        for key, value in user_data.items():
                            if key not in ("username", "created_by"):
                                setattr(existing_user, key, value)
                        existing_user.modified_by = self.actor
                        existing_user.full_clean()
                        existing_user.save()
                        
                        # Update password if specified
                        if password:
                            existing_user.set_password(password)
                            existing_user.save()
                        
                        user = existing_user
                    else:
                        # Create new user
                        user = User(**user_data)
                        user.set_password(password)
                        user.save()
                
                successes.append({
                    "row": row_num,
                    "username": username,
                    "name": f"{first_name} {last_name}".strip(),
                    "email": email,
                    "specialty": matched_specialty,
                    "password": password if not dry_run else "***",  # Don't expose passwords in dry run
                    "department": department_name if department_name else None,
                })
            except ValidationError as exc:
                failures.append({
                    "row": row_num,
                    "error": exc.message_dict if hasattr(exc, "message_dict") else str(exc),
                    "data": row
                })
                errors_triggered = True
            except Exception as exc:
                failures.append({
                    "row": row_num,
                    "error": f"Unexpected error: {str(exc)}",
                    "data": row
                })
                errors_triggered = True
        
        # Process rows
        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation
        
        # Prepare operation details
        details = {
            "successes": successes,
            "failures": failures,
        }
        
        operation.mark_completed(
            len(successes),
            len(failures),
            details,
        )
        return operation

    def import_residents(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
        generate_passwords: bool = True,
    ) -> BulkOperation:
        """
        Import resident/postgraduate data from CSV or Excel file.
        
        Expected columns (case-insensitive):
        - Name (or First Name, Last Name)
        - Year (required) - Training year (1, 2, 3, 4)
        - Specialty (required)
        - Supervisor Name (required) or Supervisor Username
        - Email (optional, will be generated)
        - Department (optional)
        - Phone (optional)
        - Registration Number (optional)
        - Username (optional, will be generated)
        - Date of Joining (optional)
        
        Creates accounts with role 'pg' and links to supervisors.
        Handles cases where supervisors don't exist (creates warning/error based on allow_partial).
        """
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_IMPORT
        )
        
        try:
            rows = list(_parse_csv_rows(uploaded_file, required_columns=None))
        except ValidationError as e:
            operation.mark_failed({"error": str(e)})
            return operation
        
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False
        unlinked_residents: List[dict] = []
        
        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            
            row_num = row.get("_row_number", "unknown")
            
            # Extract data with flexible column matching
            name = (
                row.get("name") or 
                f"{row.get('first_name', '').strip()} {row.get('last_name', '').strip()}".strip()
            )
            year = row.get("year", "").strip() or row.get("training_year", "").strip()
            specialty = row.get("specialty", "").strip()
            supervisor_name = row.get("supervisor_name", "").strip() or row.get("supervisor", "").strip()
            supervisor_username = row.get("supervisor_username", "").strip()
            email = row.get("email", "").strip()
            department_name = row.get("department", "").strip()
            phone = row.get("phone", "").strip() or row.get("phone_number", "").strip()
            registration_number = row.get("registration_number", "").strip() or row.get("reg_no", "").strip()
            username = row.get("username", "").strip()
            date_joining_str = row.get("date_joining", "").strip() or row.get("date_of_joining", "").strip()
            
            # Validate required fields
            if not name:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Name' (or 'First Name' + 'Last Name')",
                    "data": row
                })
                errors_triggered = True
                return
            
            if not year:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Year' (training year: 1, 2, 3, or 4)",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Validate year
            valid_years = [choice[0] for choice in YEAR_CHOICES]
            if year not in valid_years:
                failures.append({
                    "row": row_num,
                    "error": f"Invalid year '{year}'. Must be one of: {', '.join(valid_years)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            if not specialty:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Specialty'",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Validate specialty
            valid_specialties = [choice[0] for choice in SPECIALTY_CHOICES]
            specialty_lower = specialty.lower().replace(" ", "_")
            matched_specialty = None
            for spec_code, spec_name in SPECIALTY_CHOICES:
                if (specialty_lower == spec_code.lower() or 
                    specialty.lower() == spec_name.lower() or
                    specialty_lower in spec_name.lower() or
                    spec_name.lower() in specialty_lower):
                    matched_specialty = spec_code
                    break
            
            if not matched_specialty:
                failures.append({
                    "row": row_num,
                    "error": f"Invalid specialty '{specialty}'",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse name
            try:
                first_name, last_name = _parse_name(name)
                if not first_name and not last_name:
                    failures.append({
                        "row": row_num,
                        "error": "Invalid name format",
                        "data": row
                    })
                    errors_triggered = True
                    return
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error parsing name: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse date of joining
            date_joined = None
            if date_joining_str:
                try:
                    date_joined = _parse_date(date_joining_str)
                except Exception as e:
                    if not allow_partial:
                        failures.append({
                            "row": row_num,
                            "error": f"Invalid date format: {str(e)}",
                            "data": row
                        })
                        errors_triggered = True
                        return
                    date_joined = date.today()
            else:
                date_joined = date.today()
            
            # Get supervisor
            supervisor = None
            if supervisor_username:
                try:
                    supervisor = User.objects.filter(
                        username=supervisor_username,
                        role="supervisor"
                    ).first()
                    if not supervisor:
                        failures.append({
                            "row": row_num,
                            "error": f"Supervisor with username '{supervisor_username}' not found",
                            "data": row
                        })
                        if not allow_partial:
                            errors_triggered = True
                            return
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "error": f"Error finding supervisor by username: {str(e)}",
                        "data": row
                    })
                    if not allow_partial:
                        errors_triggered = True
                        return
            elif supervisor_name:
                try:
                    supervisor = _get_or_create_supervisor(
                        supervisor_name, 
                        self.actor,
                        specialty=matched_specialty,
                        generate_password=generate_passwords
                    )
                    if not supervisor:
                        failures.append({
                            "row": row_num,
                            "error": f"Could not create/find supervisor '{supervisor_name}'",
                            "data": row
                        })
                        if not allow_partial:
                            errors_triggered = True
                            return
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "error": f"Error with supervisor '{supervisor_name}': {str(e)}",
                        "data": row
                    })
                    if not allow_partial:
                        errors_triggered = True
                        return
            else:
                # No supervisor provided
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Supervisor Name' or 'Supervisor Username'",
                    "data": row
                })
                if not allow_partial:
                    errors_triggered = True
                    return
            
            # Generate username if not provided
            if not username:
                try:
                    username = _generate_username(first_name, last_name)
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "error": f"Error generating username: {str(e)}",
                        "data": row
                    })
                    errors_triggered = True
                    return
            
            # Generate email if not provided
            if not email:
                email = f"{username}.pgr@pmc.edu.pk"
            
            # Get or create department if provided (optional)
            department = None
            if department_name and Department:
                try:
                    department = Department.objects.filter(
                        name__iexact=department_name,
                        is_active=True
                    ).first()
                    if not department and allow_partial:
                        pass  # Continue without department
                except Exception:
                    pass  # Ignore department errors if allow_partial
            
            # Prepare user data
            user_data = {
                "username": username,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": "pg",
                "specialty": matched_specialty,
                "year": year,
                "phone_number": phone if phone else None,
                "registration_number": registration_number if registration_number else None,
                "date_joined": date_joined,
                "is_active": True,
                "created_by": self.actor,
            }
            
            if supervisor:
                user_data["supervisor"] = supervisor
            
            # Generate password
            password = None
            if generate_passwords:
                password = _generate_password_from_username(username, year)
            else:
                password = _generate_secure_password()
            
            # Create or validate user
            try:
                if dry_run:
                    # Validate without creating
                    # For dry run validation, we need a supervisor for PG role validation
                    # Use a dummy supervisor if none exists (only for validation)
                    validation_data = user_data.copy()
                    if not supervisor:
                        # Find any existing supervisor for validation purposes
                        dummy_supervisor = User.objects.filter(role="supervisor").first()
                        if dummy_supervisor:
                            validation_data["supervisor"] = dummy_supervisor
                        elif not allow_partial:
                            # If no supervisor exists at all and allow_partial=False, skip validation
                            # (This case should have been caught earlier, but handle gracefully)
                            failures.append({
                                "row": row_num,
                                "error": "No supervisor found and no supervisors exist in system for validation",
                                "data": row
                            })
                            errors_triggered = True
                            return
                    
                    user = User(**validation_data)
                    user.full_clean()
                else:
                    # Check if user already exists
                    existing_user = User.objects.filter(username=username).first()
                    if existing_user:
                        # Update existing user
                        for key, value in user_data.items():
                            if key not in ("username", "created_by"):
                                setattr(existing_user, key, value)
                        existing_user.modified_by = self.actor
                        existing_user.full_clean()
                        existing_user.save()
                        
                        # Update password if specified
                        if password:
                            existing_user.set_password(password)
                            existing_user.save()
                        
                        user = existing_user
                    else:
                        # Create new user
                        if not supervisor:
                            # Cannot create PG without supervisor (validation will fail)
                            failures.append({
                                "row": row_num,
                                "error": "Cannot create PG user without supervisor",
                                "data": row
                            })
                            errors_triggered = True
                            return
                        
                        user = User(**user_data)
                        user.set_password(password)
                        user.save()
                
                success_entry = {
                    "row": row_num,
                    "username": username,
                    "name": f"{first_name} {last_name}".strip(),
                    "email": email,
                    "specialty": matched_specialty,
                    "year": year,
                    "password": password if not dry_run else "***",
                    "supervisor": supervisor.username if supervisor else None,
                }
                
                if not supervisor:
                    success_entry["warning"] = "Created without supervisor - will need manual linking"
                    unlinked_residents.append(success_entry)
                
                successes.append(success_entry)
            except ValidationError as exc:
                failures.append({
                    "row": row_num,
                    "error": exc.message_dict if hasattr(exc, "message_dict") else str(exc),
                    "data": row
                })
                errors_triggered = True
            except Exception as exc:
                failures.append({
                    "row": row_num,
                    "error": f"Unexpected error: {str(exc)}",
                    "data": row
                })
                errors_triggered = True
        
        # Process rows
        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation
        
        # Prepare operation details
        details = {
            "successes": successes,
            "failures": failures,
            "unlinked_residents": unlinked_residents,
        }
        
        operation.mark_completed(
            len(successes),
            len(failures),
            details,
        )
        return operation

    def import_trainees(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
    ) -> BulkOperation:
        """
        Import trainee data from Excel file.
        
        Expected columns:
        - Sr. No. (ignored)
        - Name of Trainee
        - Date of Joining
        - MS/FCPS (optional)
        - Supervisor Name (optional, will create if not found)
        """
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_IMPORT
        )
        rows = list(_parse_trainee_rows(uploaded_file))
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False
        created_supervisors: List[dict] = []

        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            
            row_num = row.get("_row_number", "unknown")
            name = row.get("name", "").strip()
            date_joining = row.get("date_joining")
            qualification = row.get("qualification", "").strip()
            supervisor_name = row.get("supervisor_name", "").strip()
            
            # Validate required fields
            if not name:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Name of Trainee'",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse name
            try:
                first_name, last_name = _parse_name(name)
                if not first_name and not last_name:
                    failures.append({
                        "row": row_num,
                        "error": "Invalid name format",
                        "data": row
                    })
                    errors_triggered = True
                    return
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error parsing name: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse date
            try:
                date_joined = _parse_date(str(date_joining)) if date_joining else date.today()
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Invalid date format: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Generate username
            try:
                username = _generate_username(first_name, last_name)
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error generating username: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Infer training year
            try:
                year = _infer_training_year(date_joined)
            except Exception as e:
                year = "1"  # Default to year 1 on error
                failures.append({
                    "row": row_num,
                    "warning": f"Could not infer year, defaulting to 1: {str(e)}",
                    "data": row
                })
            
            # Get or create supervisor
            supervisor = None
            if supervisor_name:
                try:
                    supervisor = _get_or_create_supervisor(supervisor_name, self.actor)
                    if supervisor:
                        if supervisor.id not in [s["id"] for s in created_supervisors]:
                            created_supervisors.append({
                                "id": supervisor.id,
                                "username": supervisor.username,
                                "name": supervisor.get_full_name()
                            })
                    else:
                        # Supervisor creation returned None
                        if not allow_partial:
                            failures.append({
                                "row": row_num,
                                "error": f"Could not create/find supervisor '{supervisor_name}'",
                                "data": row
                            })
                            errors_triggered = True
                            return
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "warning": f"Error with supervisor '{supervisor_name}': {str(e)}",
                        "data": row
                    })
                    # Continue without supervisor if allow_partial, otherwise fail
                    if not allow_partial:
                        errors_triggered = True
                        return
            else:
                # Supervisor name not provided
                if not allow_partial:
                    failures.append({
                        "row": row_num,
                        "error": "Supervisor Name is required",
                        "data": row
                    })
                    errors_triggered = True
                    return
            
            # Generate email
            email = f"{username}.pgr@pmc.edu.pk"
            
            # Prepare user data
            user_data = {
                "username": username,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": "pg",
                "specialty": "urology",
                "year": year,
                "date_joined": date_joined,
                "registration_number": qualification if qualification else "",
                "is_active": True,
                "created_by": self.actor,
            }
            
            if supervisor:
                user_data["supervisor"] = supervisor
            
            # Create or validate user
            try:
                if dry_run:
                    # Validate without creating
                    user = User(**user_data)
                    user.full_clean()
                else:
                    # Check if user already exists
                    existing_user = User.objects.filter(username=username).first()
                    if existing_user:
                        # Update existing user
                        for key, value in user_data.items():
                            if key != "username" and key != "created_by":
                                setattr(existing_user, key, value)
                        existing_user.modified_by = self.actor
                        existing_user.full_clean()
                        existing_user.save()
                        user = existing_user
                    else:
                        # Create new user - set password before saving to avoid validation errors
                        user = User(**user_data)
                        # Set password (default password, should be changed on first login)
                        user.set_password("changeme123")
                        # Now save (this will validate and create the record)
                        user.save()
                
                successes.append({
                    "row": row_num,
                    "username": username,
                    "name": f"{first_name} {last_name}".strip(),
                    "email": email,
                    "year": year,
                    "supervisor": supervisor.username if supervisor else None,
                })
            except ValidationError as exc:
                failures.append({
                    "row": row_num,
                    "error": exc.message_dict if hasattr(exc, "message_dict") else str(exc),
                    "data": row
                })
                errors_triggered = True
            except Exception as exc:
                failures.append({
                    "row": row_num,
                    "error": f"Unexpected error: {str(exc)}",
                    "data": row
                })
                errors_triggered = True

        # Process rows
        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation

        # Prepare operation details
        details = {
            "successes": successes,
            "failures": failures,
            "created_supervisors": created_supervisors,
        }
        
        operation.mark_completed(
            len(successes),
            len(failures),
            details,
        )
        return operation


def _chunked(items: Sequence[int], chunk_size: int) -> Iterator[List[int]]:
    chunk: List[int] = []
    for item in items:
        chunk.append(item)
        if len(chunk) == chunk_size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


REQUIRED_COLUMNS = {"pg_username", "case_title", "date", "status"}


def _parse_rows(uploaded_file) -> Iterator[dict]:
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    if name.endswith(".csv"):
        text_stream = (
            io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
        )
        reader = csv.DictReader(text_stream)
        _validate_headers(reader.fieldnames)
        for row in reader:
            yield {key: (value or "").strip() for key, value in row.items() if key}
    elif name.endswith(".xlsx"):
        workbook = load_workbook(stream)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        _validate_headers(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            payload = {
                headers[idx]: (value or "") for idx, value in enumerate(row) if idx < len(headers)
            }
            yield {key: str(value).strip() for key, value in payload.items()}
    else:
        raise ValidationError("Unsupported file format")


def _validate_headers(headers: Iterable[str]) -> None:
    if not headers:
        raise ValidationError("No headers found in file")
    missing = REQUIRED_COLUMNS - set(header.strip() for header in headers if header)
    if missing:
        raise ValidationError(f"Missing columns: {', '.join(sorted(missing))}")


# ------------------------------------------------------------------
# Trainee import helper functions


def _parse_name(full_name: str) -> Tuple[str, str]:
    """
    Split full name into first_name and last_name.
    - Remove titles like Dr., Mr., Mrs., etc.
    - Last word → last_name
    - All other words → first_name
    """
    if not full_name or not full_name.strip():
        return ("", "")
    
    # Remove common titles
    title_pattern = r"^(Dr\.?|Mr\.?|Mrs\.?|Ms\.?|Prof\.?|Professor)\s+"
    name = re.sub(title_pattern, "", full_name.strip(), flags=re.IGNORECASE)
    
    # Split by spaces
    parts = [p.strip() for p in name.split() if p.strip()]
    
    if not parts:
        return ("", "")
    elif len(parts) == 1:
        return (parts[0], "")
    else:
        # Last word is last_name, rest is first_name
        return (" ".join(parts[:-1]), parts[-1])


def _generate_username(first_name: str, last_name: str) -> str:
    """
    Generate username as firstname.lastname.
    - Convert to lowercase
    - Remove special characters, keep only alphanumeric and dots
    - Check for duplicates and append number if needed
    """
    # Clean names
    first_clean = re.sub(r"[^a-z0-9]", "", first_name.lower())
    last_clean = re.sub(r"[^a-z0-9]", "", last_name.lower())
    
    if not first_clean and not last_clean:
        base_username = "trainee"
    elif not last_clean:
        base_username = first_clean
    elif not first_clean:
        base_username = last_clean
    else:
        base_username = f"{first_clean}.{last_clean}"
    
    # Check for duplicates
    username = base_username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
    
    return username


def _infer_training_year(date_joined: date) -> str:
    """
    Calculate training year based on months since date_joined.
    - 0-12 months → "1"
    - 12-24 months → "2"
    - 24-36 months → "3"
    - 36+ months → "4"
    """
    today = date.today()
    if date_joined > today:
        # Future date, default to year 1
        return "1"
    
    # Calculate months difference
    months_diff = (today.year - date_joined.year) * 12 + (today.month - date_joined.month)
    
    if months_diff < 12:
        return "1"
    elif months_diff < 24:
        return "2"
    elif months_diff < 36:
        return "3"
    else:
        return "4"


def _get_or_create_supervisor(
    supervisor_name: str, 
    actor: User, 
    specialty: Optional[str] = None,
    generate_password: bool = True
) -> Optional[User]:
    """
    Find supervisor by name or create if not found.
    - Search by full name (case-insensitive)
    - If not found, create new supervisor user
    - Uses provided specialty or defaults to urology
    - Generates username, email, and password
    
    Args:
        supervisor_name: Full name of supervisor
        actor: User creating the supervisor
        specialty: Specialty code (optional, defaults to 'urology')
        generate_password: Whether to generate a secure password
    
    Returns:
        User object (supervisor) or None if creation fails
    """
    if not supervisor_name or not supervisor_name.strip():
        return None
    
    supervisor_name = supervisor_name.strip()
    
    # Default specialty if not provided
    if not specialty:
        specialty = "urology"
    
    # Validate specialty
    valid_specialties = [choice[0] for choice in SPECIALTY_CHOICES]
    if specialty not in valid_specialties:
        specialty = "urology"  # Default fallback
    
    # Try to find existing supervisor by name
    # Search by first_name + last_name or full name match
    first_name, last_name = _parse_name(supervisor_name)
    
    if first_name and last_name:
        # Try exact match
        supervisor = User.objects.filter(
            first_name__iexact=first_name,
            last_name__iexact=last_name,
            role="supervisor"
        ).first()
        
        if supervisor:
            return supervisor
        
        # Try full name match (where first_name contains the full name)
        supervisor = User.objects.filter(
            first_name__iexact=supervisor_name,
            role="supervisor"
        ).first()
        
        if supervisor:
            return supervisor
    
    # Try to find by username if supervisor_name looks like a username
    if "." in supervisor_name or len(supervisor_name.split()) == 1:
        supervisor = User.objects.filter(
            username__iexact=supervisor_name,
            role="supervisor"
        ).first()
        if supervisor:
            return supervisor
    
    # Create new supervisor
    try:
        username = _generate_username(first_name, last_name)
        email = f"{username}.supervisor@pmc.edu.pk"
        
        # Check if username already exists
        if User.objects.filter(username=username).exists():
            # Try to find existing user with this username (might be different role)
            existing = User.objects.filter(username=username).first()
            if existing.role == "supervisor":
                return existing
            # If different role, append suffix
            counter = 1
            base_username = username
            while User.objects.filter(username=username).exists():
                username = f"{base_username}{counter}"
                counter += 1
            email = f"{username}.supervisor@pmc.edu.pk"
        
        supervisor = User.objects.create(
            username=username,
            email=email,
            first_name=first_name or supervisor_name,
            last_name=last_name,
            role="supervisor",
            specialty=specialty,
            is_active=True,
            created_by=actor,
        )
        
        # Set password
        if generate_password:
            password = _generate_password_from_username(username)
        else:
            password = _generate_secure_password()
        supervisor.set_password(password)
        supervisor.save()
        
        return supervisor
    except Exception as e:
        # Log error but return None to allow graceful handling
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error creating supervisor '{supervisor_name}': {str(e)}")
        return None


def _parse_date(date_str: str) -> date:
    """
    Parse date string in various formats.
    Supports: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY, etc.
    """
    if not date_str:
        raise ValueError("Empty date string")
    
    date_str = str(date_str).strip()
    
    # Try different date formats
    formats = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%Y.%m.%d",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    # If all formats fail, try parsing as Excel date serial number
    try:
        # Excel dates are sometimes stored as numbers
        serial = float(date_str)
        # Excel epoch is 1900-01-01, but Excel incorrectly treats 1900 as a leap year
        excel_epoch = datetime(1899, 12, 30)
        parsed_date = excel_epoch + timedelta(days=int(serial))
        return parsed_date.date()
    except (ValueError, OverflowError):
        pass
    
    raise ValueError(f"Unable to parse date: {date_str}")


def _generate_secure_password(length: int = 12) -> str:
    """
    Generate a secure random password.
    - Contains uppercase, lowercase, digits, and special characters
    - Minimum length 12 characters
    """
    if length < 12:
        length = 12
    
    # Character sets
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits
    special = "!@#$%^&*"
    
    # Ensure at least one character from each set
    password = [
        secrets.choice(lowercase),
        secrets.choice(uppercase),
        secrets.choice(digits),
        secrets.choice(special),
    ]
    
    # Fill the rest randomly
    all_chars = lowercase + uppercase + digits + special
    password.extend(secrets.choice(all_chars) for _ in range(length - 4))
    
    # Shuffle to avoid predictable patterns
    secrets.SystemRandom().shuffle(password)
    
    return "".join(password)


def _generate_password_from_username(username: str, year: Optional[str] = None) -> str:
    """
    Generate a deterministic password from username.
    Format: Username@Year! (if year provided) or Username@123!
    This makes passwords predictable for first-time login but still secure.
    """
    if year:
        return f"{username}@{year}!"
    return f"{username}@123!"


def _parse_csv_rows(uploaded_file, required_columns: Optional[set] = None) -> Iterator[dict]:
    """
    Parse CSV file and yield row dictionaries.
    Supports both CSV and Excel files.
    
    Args:
        uploaded_file: File object
        required_columns: Set of required column names (case-insensitive)
    
    Yields:
        dict: Row data with lowercase keys
    """
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    
    # Reset file pointer
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    
    # Handle CSV files
    if name.endswith(".csv"):
        text_stream = (
            io.TextIOWrapper(stream, encoding="utf-8") 
            if isinstance(stream, io.BytesIO) else stream
        )
        reader = csv.DictReader(text_stream)
        
        if reader.fieldnames:
            # Normalize headers (strip, lowercase)
            headers = [h.strip().lower() if h else f"col_{i}" 
                      for i, h in enumerate(reader.fieldnames)]
            reader.fieldnames = headers
            
            # Validate required columns
            if required_columns:
                missing = required_columns - {h.lower() for h in headers}
                if missing:
                    raise ValidationError(
                        f"Missing required columns: {', '.join(sorted(missing))}"
                    )
        
        for row_idx, row in enumerate(reader, start=2):
            # Normalize row values
            normalized_row = {
                k.lower().strip(): (v.strip() if v else "") 
                for k, v in row.items() if k
            }
            normalized_row["_row_number"] = row_idx
            yield normalized_row
    
    # Handle Excel files
    elif name.endswith((".xlsx", ".xls")):
        workbook = load_workbook(stream)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [
            str(cell.value).strip().lower() if cell.value else f"col_{idx}"
            for idx, cell in enumerate(next(sheet.iter_rows(max_row=1)))
        ]
        
        # Validate required columns
        if required_columns:
            missing = required_columns - {h.lower() for h in headers}
            if missing:
                raise ValidationError(
                    f"Missing required columns: {', '.join(sorted(missing))}"
                )
        
        # Parse data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            payload = {
                headers[idx] if idx < len(headers) else f"col_{idx}": (
                    str(value).strip() if value is not None else ""
                )
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            payload["_row_number"] = row_idx
            yield payload
    
    else:
        raise ValidationError("Unsupported file format. Please upload CSV or Excel file.")


def _parse_trainee_rows(uploaded_file) -> Iterator[dict]:
    """
    Parse Excel file with trainee data.
    Expected columns: Sr. No., Name of Trainee, Date of Joining, MS/FCPS, Supervisor Name
    """
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    
    if not name.endswith((".xlsx", ".xls")):
        raise ValidationError("Trainee import only supports Excel files (.xlsx, .xls)")
    
    workbook = load_workbook(stream)
    sheet = workbook.active
    
    # Get headers from first row
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(max_row=1))]
    
    # Map column indices
    # Expected: Sr. No., Name of Trainee, Date of Joining, MS/FCPS, Supervisor Name
    col_map = {}
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if "sr" in header_lower and "no" in header_lower:
            col_map["sr_no"] = idx
        elif "name" in header_lower and "trainee" in header_lower:
            col_map["name"] = idx
        elif "date" in header_lower and "joining" in header_lower:
            col_map["date"] = idx
        elif "ms" in header_lower or "fcps" in header_lower:
            col_map["qualification"] = idx
        elif "supervisor" in header_lower and "name" in header_lower:
            col_map["supervisor"] = idx
    
    # Validate required columns
    if "name" not in col_map or "date" not in col_map:
        raise ValidationError("Missing required columns: 'Name of Trainee' and 'Date of Joining'")
    
    # Parse data rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        row_data = {}
        if "name" in col_map and col_map["name"] < len(row):
            row_data["name"] = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
        if "date" in col_map and col_map["date"] < len(row):
            row_data["date_joining"] = row[col_map["date"]]
        if "qualification" in col_map and col_map["qualification"] < len(row):
            row_data["qualification"] = str(row[col_map["qualification"]]).strip() if row[col_map["qualification"]] else ""
        if "supervisor" in col_map and col_map["supervisor"] < len(row):
            row_data["supervisor_name"] = str(row[col_map["supervisor"]]).strip() if row[col_map["supervisor"]] else ""
        
        row_data["_row_number"] = row_idx
        yield row_data


def generate_trainee_template() -> io.BytesIO:
    """
    Generate a template Excel file for trainee import.
    Returns a BytesIO object containing the Excel file.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trainee Data"
    
    # Set headers
    headers = ["Sr. No.", "Name of Trainee", "Date of Joining", "MS/FCPS", "Supervisor Name"]
    header_font = Font(bold=True)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
    
    # Add example rows
    examples = [
        [1, "John Doe", "2024-01-15", "MS", "Dr. Smith"],
        [2, "Jane Smith", "2024-02-01", "FCPS", "Dr. Johnson"],
        [3, "Ahmed Ali", "2024-03-10", "", "Dr. Khan"],
    ]
    
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, value in enumerate(example, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        column_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[column_letter].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def convert_excel_to_trainee_format(uploaded_file) -> io.BytesIO:
    """
    Convert any Excel file to the required trainee import format.
    Attempts to intelligently map columns from the uploaded file to the required format.
    Returns a BytesIO object containing the converted Excel file.
    """
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    
    if not name.endswith((".xlsx", ".xls")):
        raise ValidationError("Only Excel files (.xlsx, .xls) are supported")
    
    # Load the uploaded workbook
    source_workbook = load_workbook(stream)
    source_sheet = source_workbook.active
    
    # Get headers from first row
    source_headers = []
    for cell in next(source_sheet.iter_rows(max_row=1)):
        header = str(cell.value).strip() if cell.value else ""
        source_headers.append(header)
    
    # Create new workbook with required format
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Trainee Data"
    
    # Required headers
    required_headers = ["Sr. No.", "Name of Trainee", "Date of Joining", "MS/FCPS", "Supervisor Name"]
    header_font = Font(bold=True)
    
    # Write headers
    for col_idx, header in enumerate(required_headers, start=1):
        cell = output_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
    
    # Try to map source columns to required columns
    column_mapping = {}
    for req_col in required_headers:
        req_lower = req_col.lower()
        best_match_idx = None
        best_match_score = 0
        
        for src_idx, src_header in enumerate(source_headers):
            if not src_header:
                continue
            src_lower = src_header.lower()
            
            # Calculate match score
            score = 0
            if req_col == "Sr. No.":
                if any(keyword in src_lower for keyword in ["sr", "serial", "no", "number", "sno"]):
                    score = 1
            elif req_col == "Name of Trainee":
                if any(keyword in src_lower for keyword in ["name", "trainee", "student", "pg"]):
                    score = 1
                    if "trainee" in src_lower or "name" in src_lower:
                        score = 2
            elif req_col == "Date of Joining":
                if any(keyword in src_lower for keyword in ["date", "joining", "join", "doj"]):
                    score = 1
                    if "joining" in src_lower or "doj" in src_lower:
                        score = 2
            elif req_col == "MS/FCPS":
                if any(keyword in src_lower for keyword in ["ms", "fcps", "qualification", "degree"]):
                    score = 1
            elif req_col == "Supervisor Name":
                if any(keyword in src_lower for keyword in ["supervisor", "mentor", "guide"]):
                    score = 1
                    if "supervisor" in src_lower:
                        score = 2
            
            if score > best_match_score:
                best_match_score = score
                best_match_idx = src_idx
        
        if best_match_idx is not None:
            column_mapping[req_col] = best_match_idx
    
    # Copy data rows
    output_row = 2
    for source_row_idx, source_row in enumerate(source_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(source_row):  # Skip empty rows
            continue
        
        # Map columns
        for col_idx, req_header in enumerate(required_headers, start=1):
            if req_header in column_mapping:
                src_col_idx = column_mapping[req_header]
                if src_col_idx < len(source_row):
                    value = source_row[src_col_idx]
                    # Format date if it's a date object
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d")
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")
                    output_sheet.cell(row=output_row, column=col_idx, value=value)
            elif req_header == "Sr. No.":
                # Auto-generate serial number
                output_sheet.cell(row=output_row, column=col_idx, value=output_row - 1)
        
        output_row += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, len(required_headers) + 1):
        column_letter = output_sheet.cell(row=1, column=col_idx).column_letter
        output_sheet.column_dimensions[column_letter].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    output_workbook.save(output)
    output.seek(0)
    return output
