"""API views for executing bulk operations."""

from __future__ import annotations

import json
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import ValidationError as DjangoValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import permissions, serializers, status, viewsets
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView
from drf_spectacular.utils import extend_schema

from sims.bulk.models import BulkOperation, MappingPreset, FlexibleImportAudit
from sims.bulk.serializers import (
    BulkAssignmentSerializer,
    BulkImportSerializer,
    BulkReviewSerializer,
    DepartmentImportSerializer,
    ResidentImportSerializer,
    SupervisorImportSerializer,
    TraineeImportSerializer,
    MappingPresetSerializer,
)
from sims.bulk.services import BulkService

User = get_user_model()


class BulkEmptySchemaSerializer(serializers.Serializer):
    pass


def _operation_payload(operation: BulkOperation) -> dict:
    return {
        "operation": operation.operation,
        "status": operation.status,
        "success_count": operation.success_count,
        "failure_count": operation.failure_count,
        "details": operation.details,
        "created_at": operation.created_at,
        "completed_at": operation.completed_at,
    }


def _track_bulk_event(request, *, event_type, resource, operation=None, error_code=None):
    pass  # analytics module removed


@extend_schema(responses={200: None})
class BulkReviewView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        serializer = BulkReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        service = BulkService(request.user)
        operation = service.review_entries(**serializer.validated_data)
        return Response(_operation_payload(operation), status=status.HTTP_200_OK)


@extend_schema(responses={200: None})
class BulkAssignmentView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        serializer = BulkAssignmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        supervisor = get_object_or_404(
            User.objects.filter(role="supervisor"),
            pk=serializer.validated_data["supervisor_id"],
        )
        service = BulkService(request.user)
        operation = service.assign_supervisor(
            entry_ids=serializer.validated_data["entry_ids"], supervisor=supervisor
        )
        return Response(_operation_payload(operation), status=status.HTTP_200_OK)


@extend_schema(responses={200: None})
class BulkImportView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(operation_id="api_bulk_logbook_import_create")
    def post(self, request: Request) -> Response:
        from django.core.exceptions import ValidationError as DjangoValidationError

        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploaded_file = serializer.validated_data["file"]
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.import.started", resource="logbook")

        try:
            operation = service.import_logbook_entries(
                uploaded_file,
                dry_run=serializer.validated_data["dry_run"],
                allow_partial=serializer.validated_data["allow_partial"],
            )
        except DjangoValidationError as e:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="logbook",
                error_code="validation_error",
            )
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if operation.status == BulkOperation.STATUS_COMPLETED:
            _track_bulk_event(
                request,
                event_type="data.import.completed",
                resource="logbook",
                operation=operation,
            )
        else:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="logbook",
                operation=operation,
                error_code="operation_failed",
            )

        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(_operation_payload(operation), status=status_code)


@extend_schema(responses={200: None})
class BulkTraineeImportView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        from django.core.exceptions import ValidationError as DjangoValidationError

        serializer = TraineeImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploaded_file = serializer.validated_data["file"]
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.import.started", resource="trainees")

        try:
            operation = service.import_trainees(
                uploaded_file,
                dry_run=serializer.validated_data["dry_run"],
                allow_partial=serializer.validated_data["allow_partial"],
            )
        except DjangoValidationError as e:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="trainees",
                error_code="validation_error",
            )
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if operation.status == BulkOperation.STATUS_COMPLETED:
            _track_bulk_event(
                request,
                event_type="data.import.completed",
                resource="trainees",
                operation=operation,
            )
        else:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="trainees",
                operation=operation,
                error_code="operation_failed",
            )

        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(_operation_payload(operation), status=status_code)


@extend_schema(responses={200: None})
class BulkSupervisorImportView(APIView):
    """
    Bulk import view for supervisors/faculty.

    Accepts CSV or Excel files with supervisor data.
    Creates accounts with role 'supervisor' and generates passwords.
    """
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        from django.core.exceptions import ValidationError as DjangoValidationError

        serializer = SupervisorImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploaded_file = serializer.validated_data["file"]
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.import.started", resource="supervisors")

        try:
            operation = service.import_supervisors(
                uploaded_file,
                dry_run=serializer.validated_data["dry_run"],
                allow_partial=serializer.validated_data["allow_partial"],
                generate_passwords=serializer.validated_data.get("generate_passwords", True),
            )
        except DjangoValidationError as e:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="supervisors",
                error_code="validation_error",
            )
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if operation.status == BulkOperation.STATUS_COMPLETED:
            _track_bulk_event(
                request,
                event_type="data.import.completed",
                resource="supervisors",
                operation=operation,
            )
        else:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="supervisors",
                operation=operation,
                error_code="operation_failed",
            )

        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(_operation_payload(operation), status=status_code)


@extend_schema(responses={200: None})
class BulkResidentImportView(APIView):
    """
    Bulk import view for residents/postgraduates.

    Accepts CSV or Excel files with resident data.
    Creates accounts with role 'pg' and links to supervisors.
    Handles cases where supervisors don't exist (based on allow_partial setting).
    """
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        from django.core.exceptions import ValidationError as DjangoValidationError

        serializer = ResidentImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploaded_file = serializer.validated_data["file"]
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.import.started", resource="residents")

        try:
            operation = service.import_residents(
                uploaded_file,
                dry_run=serializer.validated_data["dry_run"],
                allow_partial=serializer.validated_data["allow_partial"],
                generate_passwords=serializer.validated_data.get("generate_passwords", True),
            )
        except DjangoValidationError as e:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="residents",
                error_code="validation_error",
            )
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        if operation.status == BulkOperation.STATUS_COMPLETED:
            _track_bulk_event(
                request,
                event_type="data.import.completed",
                resource="residents",
                operation=operation,
            )
        else:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="residents",
                operation=operation,
                error_code="operation_failed",
            )

        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(_operation_payload(operation), status=status_code)


@extend_schema(responses={200: None})
class BulkDepartmentImportView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        if getattr(request.user, "role", None) != "admin":
            return Response({"detail": "Only admins can import departments."}, status=403)

        serializer = DepartmentImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.import.started", resource="departments")
        operation = service.import_departments(
            serializer.validated_data["file"],
            dry_run=serializer.validated_data["dry_run"],
            allow_partial=serializer.validated_data["allow_partial"],
        )
        if operation.status == BulkOperation.STATUS_COMPLETED:
            _track_bulk_event(
                request,
                event_type="data.import.completed",
                resource="departments",
                operation=operation,
            )
        else:
            _track_bulk_event(
                request,
                event_type="data.import.failed",
                resource="departments",
                operation=operation,
                error_code="operation_failed",
            )
        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(_operation_payload(operation), status=status_code)


@extend_schema(responses={200: None})
class BulkExportView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request: Request, resource: str) -> HttpResponse:
        from django.core.exceptions import ValidationError as DjangoValidationError

        if getattr(request.user, "role", None) not in {"admin", "utrmc_admin"}:
            return Response({"detail": "Only admins can export bulk datasets."}, status=403)
        export_format = request.query_params.get("file_format", "xlsx").lower()
        service = BulkService(request.user)
        _track_bulk_event(request, event_type="data.export.started", resource=resource)
        try:
            export_file = service.export_dataset(resource=resource, export_format=export_format)
        except DjangoValidationError as exc:
            _track_bulk_event(
                request,
                event_type="data.export.failed",
                resource=resource,
                error_code="validation_error",
            )
            return Response({"detail": str(exc)}, status=400)
        _track_bulk_event(request, event_type="data.export.completed", resource=resource)
        response = HttpResponse(export_file.content, content_type=export_file.content_type)
        response["Content-Disposition"] = f'attachment; filename="{export_file.filename}"'
        return response


@extend_schema(responses={200: None})
class BulkTemplateView(APIView):
    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request: Request, resource: str) -> HttpResponse:
        from django.core.exceptions import ValidationError as DjangoValidationError

        if getattr(request.user, "role", None) not in {"admin", "utrmc_admin"}:
            return Response({"detail": "Only admins can download bulk templates."}, status=403)
        service = BulkService(request.user)
        try:
            export_file = service.export_template(resource)
        except DjangoValidationError as exc:
            return Response({"detail": str(exc)}, status=400)
        response = HttpResponse(export_file.content, content_type=export_file.content_type)
        response["Content-Disposition"] = f'attachment; filename="{export_file.filename}"'
        return response


# ---------------------------------------------------------------------------
# NEW unified import endpoint: /api/bulk/import/<entity>/<action>/
# action = "dry-run" | "apply"
# entity = hospitals | matrix | supervision-links | departments | supervisors | residents | trainees

_ENTITY_METHOD_MAP = {
    "hospitals": "import_userbase_hospitals",
    "matrix": "import_userbase_matrix",
    "supervision-links": "import_userbase_supervision_links",
    "departments": "import_userbase_departments",
    "faculty-supervisors": "import_userbase_faculty_supervisors",
    "residents": "import_userbase_residents",
    "hod-assignments": "import_userbase_hod_assignments",
    "rotation-assignments": "import_userbase_rotation_assignments",
    "supervisors": "import_supervisors",
    "trainees": "import_trainees",
    # Training module
    "training-programs": "import_training_programs",
    "rotation-templates": "import_rotation_templates",
    "resident-training-records": "import_resident_training_records",
}

_ALLOWED_ROLES = {"admin", "utrmc_admin"}


@extend_schema(responses={200: None})
class BulkImportEntityView(APIView):
    """Unified import endpoint.

    POST /api/bulk/import/<entity>/dry-run/   — validate only
    POST /api/bulk/import/<entity>/apply/     — write to DB
    """

    serializer_class = BulkEmptySchemaSerializer
    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(operation_id="api_bulk_entity_import_action_create")
    def post(self, request: Request, entity: str, action: str) -> Response:
        from django.core.exceptions import ValidationError as DjangoValidationError

        if getattr(request.user, "role", None) not in _ALLOWED_ROLES:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        if action not in {"dry-run", "apply"}:
            return Response({"detail": "action must be 'dry-run' or 'apply'."}, status=status.HTTP_400_BAD_REQUEST)

        method_name = _ENTITY_METHOD_MAP.get(entity)
        if not method_name:
            return Response({"detail": f"Unknown entity '{entity}'."}, status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = (action == "dry-run")
        try:
            service = BulkService(request.user)
        except Exception as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_403_FORBIDDEN)

        _track_bulk_event(request, event_type="data.import.started", resource=entity)
        try:
            operation = getattr(service, method_name)(
                uploaded_file,
                dry_run=dry_run,
                allow_partial=True,
            )
        except DjangoValidationError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload = _operation_payload(operation)
        payload["dry_run"] = dry_run
        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(payload, status=status_code)


        status_code = (
            status.HTTP_200_OK
            if operation.status == BulkOperation.STATUS_COMPLETED
            else status.HTTP_400_BAD_REQUEST
        )
        return Response(payload, status=status_code)


# ---------------------------------------------------------------------------
# FLEXIBLE COLUMN MAPPING IMPORT VIEWS
# ---------------------------------------------------------------------------

FLEXIBLE_SCHEMAS = {
    "residents": {
        "label": "Residents",
        "fields": [
            {"name": "email", "label": "Email", "required": True, "type": "email"},
            {"name": "full_name", "label": "Full Name", "required": True, "type": "string"},
            {"name": "phone_number", "label": "Phone Number", "required": False, "type": "string"},
            {"name": "role", "label": "Role (resident/pg)", "required": False, "type": "choice", "choices": ["resident", "pg"]},
            {"name": "specialty", "label": "Specialty", "required": True, "type": "string"},
            {"name": "year", "label": "Year (1, 2, 3, 4)", "required": True, "type": "string"},
            {"name": "pgr_id", "label": "PGR ID", "required": False, "type": "string"},
            {"name": "training_start", "label": "Training Start Date (YYYY-MM-DD)", "required": True, "type": "date"},
            {"name": "training_end", "label": "Training End Date (YYYY-MM-DD)", "required": False, "type": "date"},
            {"name": "training_level", "label": "Training Level (e.g. Y1)", "required": False, "type": "string"},
            {"name": "department_code", "label": "Department Code", "required": False, "type": "string"},
            {"name": "hospital_code", "label": "Hospital Code", "required": False, "type": "string"},
            {"name": "supervisor_email", "label": "Supervisor Email", "required": False, "type": "email"},
            {"name": "username", "label": "Username", "required": False, "type": "string"},
            {"name": "password", "label": "Password", "required": False, "type": "string"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "faculty-supervisors": {
        "label": "Supervisors",
        "fields": [
            {"name": "email", "label": "Email", "required": True, "type": "email"},
            {"name": "full_name", "label": "Full Name", "required": True, "type": "string"},
            {"name": "phone_number", "label": "Phone Number", "required": False, "type": "string"},
            {"name": "role", "label": "Role (faculty/supervisor)", "required": True, "type": "choice", "choices": ["faculty", "supervisor"]},
            {"name": "specialty", "label": "Specialty", "required": False, "type": "string"},
            {"name": "department_code", "label": "Department Code", "required": False, "type": "string"},
            {"name": "hospital_code", "label": "Hospital Code", "required": False, "type": "string"},
            {"name": "designation", "label": "Designation", "required": False, "type": "string"},
            {"name": "registration_number", "label": "Registration/PMDC Number", "required": False, "type": "string"},
            {"name": "username", "label": "Username", "required": False, "type": "string"},
            {"name": "password", "label": "Password", "required": False, "type": "string"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
            {"name": "start_date", "label": "Start Date (YYYY-MM-DD)", "required": False, "type": "date"},
        ]
    },
    "rotation-assignments": {
        "label": "Resident Placement",
        "fields": [
            {"name": "resident_email", "label": "Resident Email", "required": True, "type": "email"},
            {"name": "hospital_code", "label": "Hospital Code", "required": True, "type": "string"},
            {"name": "department_code", "label": "Department Code", "required": True, "type": "string"},
            {"name": "start_date", "label": "Start Date (YYYY-MM-DD)", "required": True, "type": "date"},
            {"name": "end_date", "label": "End Date (YYYY-MM-DD)", "required": True, "type": "date"},
            {"name": "status", "label": "Status (DRAFT/SUBMITTED/APPROVED/ACTIVE/COMPLETED)", "required": False, "type": "string"},
            {"name": "notes", "label": "Notes", "required": False, "type": "string"},
        ]
    },
    "supervision-links": {
        "label": "Supervisor Assignment",
        "fields": [
            {"name": "supervisor_email", "label": "Supervisor Email", "required": True, "type": "email"},
            {"name": "resident_email", "label": "Resident Email", "required": True, "type": "email"},
            {"name": "department_code", "label": "Department Code", "required": False, "type": "string"},
            {"name": "start_date", "label": "Start Date (YYYY-MM-DD)", "required": True, "type": "date"},
            {"name": "end_date", "label": "End Date (YYYY-MM-DD)", "required": False, "type": "date"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "hospitals": {
        "label": "Hospitals",
        "fields": [
            {"name": "hospital_code", "label": "Hospital Code", "required": True, "type": "string"},
            {"name": "hospital_name", "label": "Hospital Name", "required": True, "type": "string"},
            {"name": "address", "label": "Address", "required": False, "type": "string"},
            {"name": "phone", "label": "Phone", "required": False, "type": "string"},
            {"name": "email", "label": "Email", "required": False, "type": "email"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "departments": {
        "label": "Departments",
        "fields": [
            {"name": "department_code", "label": "Department Code", "required": True, "type": "string"},
            {"name": "department_name", "label": "Department Name", "required": True, "type": "string"},
            {"name": "description", "label": "Description", "required": False, "type": "string"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "matrix": {
        "label": "HospitalDepartment matrix",
        "fields": [
            {"name": "hospital_code", "label": "Hospital Code", "required": True, "type": "string"},
            {"name": "department_code", "label": "Department Code", "required": True, "type": "string"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "hod-assignments": {
        "label": "HODs",
        "fields": [
            {"name": "department_code", "label": "Department Code", "required": True, "type": "string"},
            {"name": "hod_email", "label": "HOD Email", "required": True, "type": "email"},
            {"name": "start_date", "label": "Start Date (YYYY-MM-DD)", "required": True, "type": "date"},
            {"name": "end_date", "label": "End Date (YYYY-MM-DD)", "required": False, "type": "date"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    },
    "rotation-templates": {
        "label": "Rotations",
        "fields": [
            {"name": "program_code", "label": "Program Code", "required": True, "type": "string"},
            {"name": "template_name", "label": "Template Name", "required": True, "type": "string"},
            {"name": "department_code", "label": "Department Code", "required": True, "type": "string"},
            {"name": "duration_weeks", "label": "Duration (Weeks)", "required": False, "type": "integer"},
            {"name": "required", "label": "Required (true/false)", "required": False, "type": "boolean"},
            {"name": "active", "label": "Active (true/false)", "required": False, "type": "boolean"},
        ]
    }
}


def _transform_custom_file_to_standard_csv(uploaded_file, entity, mapping_dict, sheet_name=None):
    import io
    import csv
    from openpyxl import load_workbook

    # 1. Parse uploaded_file into list of rows (dicts)
    uploaded_file.seek(0)
    name = getattr(uploaded_file, "name", "uploaded").lower()
    content = uploaded_file.read()
    stream = io.BytesIO(content) if isinstance(content, bytes) else io.StringIO(content)
    stream.seek(0)

    rows = []
    if name.endswith(".csv"):
        text_stream = io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
        reader = csv.DictReader(text_stream)
        rows = list(reader)
    elif name.endswith((".xlsx", ".xls")):
        workbook = load_workbook(stream, read_only=True)
        if not sheet_name or sheet_name not in workbook.sheetnames:
            sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]

        row_iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iterator)
            headers = [str(cell) if cell is not None else f"col_{idx}" for idx, cell in enumerate(header_row)]
        except StopIteration:
            headers = []

        for row in row_iterator:
            if not any(cell is not None for cell in row):
                continue
            payload = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    payload[headers[idx]] = str(val).strip() if val is not None else ""
            rows.append(payload)
    else:
        raise DjangoValidationError("Unsupported file format.")

    # 2. Map and transform rows to target schema keys
    schema = FLEXIBLE_SCHEMAS[entity]
    target_fields = [f["name"] for f in schema["fields"]]

    transformed_rows = []
    for row in rows:
        transformed_row = {}
        for field in target_fields:
            custom_col = mapping_dict.get(field)
            val = ""
            if custom_col and custom_col in row:
                val = str(row[custom_col]).strip()
            transformed_row[field] = val
        transformed_rows.append(transformed_row)

    # 3. Write transformed rows to in-memory CSV
    out_stream = io.StringIO()
    writer = csv.DictWriter(out_stream, fieldnames=target_fields)
    writer.writeheader()
    for row in transformed_rows:
        writer.writerow(row)

    out_content = out_stream.getvalue().encode("utf-8")

    from django.core.files.uploadedfile import SimpleUploadedFile
    return SimpleUploadedFile(f"{entity}_transformed.csv", out_content, content_type="text/csv")


@extend_schema(responses={200: None})
class FlexibleSchemasView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request: Request) -> Response:
        return Response(FLEXIBLE_SCHEMAS)


@extend_schema(responses={200: None})
class FlexibleDetectHeadersView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        if getattr(request.user, "role", None) not in _ALLOWED_ROLES:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=400)

        name = getattr(uploaded_file, "name", "uploaded").lower()
        sheet_name = request.data.get("sheet_name")

        import io
        content = uploaded_file.read()
        stream = io.BytesIO(content) if isinstance(content, bytes) else io.StringIO(content)
        stream.seek(0)

        headers = []
        sample_rows = []
        total_rows = 0
        sheets = []

        try:
            if name.endswith(".csv"):
                import csv
                text_stream = io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
                reader = csv.DictReader(text_stream)
                headers = list(reader.fieldnames or [])
                rows = list(reader)
                total_rows = len(rows)
                sample_rows = rows[:10]
            elif name.endswith((".xlsx", ".xls")):
                from openpyxl import load_workbook
                workbook = load_workbook(stream, read_only=True)
                sheets = workbook.sheetnames
                if not sheet_name or sheet_name not in sheets:
                    sheet_name = sheets[0]
                sheet = workbook[sheet_name]

                row_iterator = sheet.iter_rows(values_only=True)
                try:
                    header_row = next(row_iterator)
                    headers = [str(cell) if cell is not None else f"col_{idx}" for idx, cell in enumerate(header_row)]
                except StopIteration:
                    headers = []

                rows = []
                for row in row_iterator:
                    if not any(cell is not None for cell in row):
                        continue
                    payload = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            payload[headers[idx]] = str(val).strip() if val is not None else ""
                    rows.append(payload)

                total_rows = len(rows)
                sample_rows = rows[:10]
            else:
                return Response({"detail": "Unsupported file format. Upload CSV or Excel."}, status=400)
        except Exception as e:
            return Response({"detail": f"Failed to parse file: {str(e)}"}, status=400)

        return Response({
            "headers": headers,
            "sample_rows": sample_rows,
            "total_rows": total_rows,
            "sheets": sheets
        })


@extend_schema(responses={200: None})
class FlexibleValidateMappingView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        if getattr(request.user, "role", None) not in _ALLOWED_ROLES:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        entity = request.data.get("entity")
        mapping = request.data.get("mapping")

        if not entity or entity not in FLEXIBLE_SCHEMAS:
            return Response({"detail": f"Invalid or missing entity '{entity}'."}, status=400)
        if not isinstance(mapping, dict):
            return Response({"detail": "mapping must be a dictionary."}, status=400)

        schema = FLEXIBLE_SCHEMAS[entity]
        required_fields = [f["name"] for f in schema["fields"] if f["required"]]
        optional_fields = [f["name"] for f in schema["fields"] if not f["required"]]

        missing_required = []
        for req in required_fields:
            if req not in mapping or not mapping[req]:
                missing_required.append(req)

        mapped_values = [v for k, v in mapping.items() if v]
        duplicates = {}
        for val in set(mapped_values):
            occurs = [k for k, v in mapping.items() if v == val]
            if len(occurs) > 1:
                duplicates[val] = occurs

        ready = len(missing_required) == 0 and len(duplicates) == 0

        return Response({
            "ready": ready,
            "missing_required": missing_required,
            "duplicate_mappings": duplicates,
            "required_fields": required_fields,
            "optional_fields": optional_fields
        })


@extend_schema(responses={200: None})
class FlexibleDryRunView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        if getattr(request.user, "role", None) not in _ALLOWED_ROLES:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        entity = request.data.get("entity")
        mapping_str = request.data.get("mapping")
        sheet_name = request.data.get("sheet_name")
        uploaded_file = request.FILES.get("file")

        if not entity or entity not in FLEXIBLE_SCHEMAS:
            return Response({"detail": f"Invalid or missing entity '{entity}'."}, status=400)
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            mapping = json.loads(mapping_str) if isinstance(mapping_str, str) else mapping_str
        except Exception:
            return Response({"detail": "Invalid mapping JSON format."}, status=400)

        try:
            transformed_file = _transform_custom_file_to_standard_csv(uploaded_file, entity, mapping, sheet_name)
        except DjangoValidationError as e:
            return Response({"detail": str(e)}, status=400)
        except Exception as e:
            return Response({"detail": f"Failed to transform file: {str(e)}"}, status=400)

        method_name = _ENTITY_METHOD_MAP.get(entity)
        if not method_name:
            return Response({"detail": f"Unknown entity '{entity}'."}, status=400)

        try:
            service = BulkService(request.user)
            operation = getattr(service, method_name)(
                transformed_file,
                dry_run=True,
                allow_partial=True,
            )
        except Exception as exc:
            return Response({"detail": str(exc)}, status=400)

        transformed_file.seek(0)
        import csv
        import io
        text_stream = io.StringIO(transformed_file.read().decode("utf-8"))
        reader = csv.DictReader(text_stream)
        transformed_rows = list(reader)
        for idx, row in enumerate(transformed_rows, start=2):
            row["_row_number"] = idx

        payload = _operation_payload(operation)
        payload["dry_run"] = True
        payload["rows"] = transformed_rows

        return Response(payload, status=status.HTTP_200_OK)


@extend_schema(responses={200: None})
class FlexibleImportApplyView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request: Request) -> Response:
        if getattr(request.user, "role", None) not in _ALLOWED_ROLES:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        entity = request.data.get("entity")
        mapping_str = request.data.get("mapping")
        sheet_name = request.data.get("sheet_name")
        import_mode = request.data.get("import_mode", "strict")
        uploaded_file = request.FILES.get("file")

        if not entity or entity not in FLEXIBLE_SCHEMAS:
            return Response({"detail": f"Invalid or missing entity '{entity}'."}, status=400)
        if not uploaded_file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            mapping = json.loads(mapping_str) if isinstance(mapping_str, str) else mapping_str
        except Exception:
            return Response({"detail": "Invalid mapping JSON format."}, status=400)

        try:
            transformed_file = _transform_custom_file_to_standard_csv(uploaded_file, entity, mapping, sheet_name)
        except DjangoValidationError as e:
            return Response({"detail": str(e)}, status=400)
        except Exception as e:
            return Response({"detail": f"Failed to transform file: {str(e)}"}, status=400)

        method_name = _ENTITY_METHOD_MAP.get(entity)
        if not method_name:
            return Response({"detail": f"Unknown entity '{entity}'."}, status=400)

        audit = FlexibleImportAudit.objects.create(
            file_name=getattr(uploaded_file, "name", "uploaded"),
            import_type=entity,
            uploaded_by=request.user,
            mapping_used=mapping,
            status="pending"
        )

        try:
            service = BulkService(request.user)

            if import_mode == "strict":
                transformed_file.seek(0)
                dry_run_op = getattr(service, method_name)(
                    transformed_file,
                    dry_run=True,
                    allow_partial=True,
                )
                if dry_run_op.failure_count > 0:
                    audit.status = "failed"
                    audit.dry_run_result = dry_run_op.details
                    audit.failure_count = dry_run_op.failure_count
                    audit.save()
                    return Response({
                        "detail": "Strict mode: Import rejected due to row-level validation errors.",
                        "success_count": 0,
                        "failure_count": dry_run_op.failure_count,
                        "details": dry_run_op.details
                    }, status=status.HTTP_400_BAD_REQUEST)

                transformed_file.seek(0)
                with transaction.atomic():
                    operation = getattr(service, method_name)(
                        transformed_file,
                        dry_run=False,
                        allow_partial=False,
                    )
            else:
                transformed_file.seek(0)
                with transaction.atomic():
                    operation = getattr(service, method_name)(
                        transformed_file,
                        dry_run=False,
                        allow_partial=True,
                    )

            if operation.status == BulkOperation.STATUS_FAILED:
                audit.status = "failed"
                audit.final_import_result = operation.details
                audit.failure_count = operation.failure_count
                audit.save()
                return Response({
                    "detail": "Import operation failed.",
                    "success_count": 0,
                    "failure_count": operation.failure_count,
                    "details": operation.details
                }, status=status.HTTP_400_BAD_REQUEST)

            preset_id = request.data.get("preset_id")
            if preset_id:
                MappingPreset.objects.filter(pk=preset_id, created_by=request.user).update(last_used_at=timezone.now())

            audit.status = "completed"
            audit.success_count = operation.success_count
            audit.failure_count = operation.failure_count
            audit.final_import_result = operation.details
            audit.save()

            payload = _operation_payload(operation)
            payload["dry_run"] = False
            return Response(payload, status=status.HTTP_200_OK)

        except Exception as exc:
            audit.status = "failed"
            audit.final_import_result = {"error": str(exc)}
            audit.save()
            return Response({"detail": str(exc)}, status=400)


class MappingPresetViewSet(viewsets.ModelViewSet):
    serializer_class = MappingPresetSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = MappingPreset.objects.filter(created_by=self.request.user)
        entity = self.request.query_params.get("entity")
        if entity:
            queryset = queryset.filter(entity=entity)
        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


__all__ = [
    "BulkReviewView",
    "BulkAssignmentView",
    "BulkImportView",
    "BulkTraineeImportView",
    "BulkSupervisorImportView",
    "BulkResidentImportView",
    "BulkDepartmentImportView",
    "BulkExportView",
    "BulkTemplateView",
    "BulkImportEntityView",
    "FlexibleSchemasView",
    "FlexibleDetectHeadersView",
    "FlexibleValidateMappingView",
    "FlexibleDryRunView",
    "FlexibleImportApplyView",
    "MappingPresetViewSet",
]
