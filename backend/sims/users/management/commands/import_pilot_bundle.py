from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from sims.academics.models import Department
from sims.bulk.services import BulkService
from sims.bulk.userbase_engine import import_entity
from sims.rotations.models import Hospital

User = get_user_model()

SEARCH_ROOTS = [
    Path("/home/munaim/srv/apps/pgsims/pilot_data/first_pilot_run"),
    Path("/home/munaim/srv/apps/pgsims/pilot_data"),
    Path("/home/munaim/srv/apps/pgsims"),
    Path("/home/munaim/srv"),
]
SPREADSHEET_SUFFIXES = {".csv", ".xls", ".xlsx", ".xlsm", ".ods"}
DISCOVERY_PATTERNS = {
    "workbook": ["*pilot*workbook*", "*pilot*", "*final*workbook*", "*workbook*"],
    "supervisors": ["*final*supervisor*", "*pilot*supervisor*", "*supervisor*list*"],
    "residents": ["*final*resident*", "*pilot*resident*", "*resident*list*"],
}

SHEET_HINTS = {
    "supervisors": {"supervisor", "faculty", "consultant"},
    "residents": {"resident", "trainee", "pgr", "pg"},
    "links": {"supervision", "link", "assignment"},
    "programs": {"program"},
    "training_records": {"training", "enrollment"},
}

ALIASES = {
    "supervisors": {
        "email": ["email", "supervisor_email"],
        "full_name": ["full_name", "name", "supervisor_name", "supervisor"],
        "phone_number": ["phone", "phone_number", "mobile", "contact", "contact_number"],
        "designation": ["designation", "title", "role_title"],
        "registration_number": ["registration_number", "pmc_no", "pmdc_no", "license_number"],
        "department_code": ["department_code", "department", "dept", "specialty_department"],
        "hospital_code": ["hospital_code", "hospital", "site", "hospital_name"],
        "specialty": ["specialty", "discipline"],
        "start_date": ["start_date", "joining_date", "date_of_joining", "date_joining"],
        "role": ["role"],
    },
    "residents": {
        "email": ["email", "resident_email", "trainee_email"],
        "full_name": ["full_name", "name", "resident_name", "trainee_name", "name_of_trainee"],
        "phone_number": ["phone", "phone_number", "mobile", "contact", "contact_number"],
        "year": ["year", "training_year", "training year", "current_year", "yr"],
        "pgr_id": ["pgr_id", "registration_number", "roll_number", "resident_id", "trainee_id"],
        "training_start": ["training_start", "start_date", "date_joining", "date_of_joining", "joining_date"],
        "training_end": ["training_end", "expected_end_date", "end_date"],
        "training_level": ["training_level", "level", "current_level"],
        "department_code": ["department_code", "department", "dept"],
        "hospital_code": ["hospital_code", "hospital", "site", "hospital_name"],
        "specialty": ["specialty", "discipline"],
        "supervisor_email": ["supervisor_email"],
        "supervisor_name": ["supervisor_name", "supervisor"],
        "role": ["role"],
        "program_track": ["track", "program", "qualification", "ms_fcps", "ms/fcps"],
        "thesis_status": ["thesis_status", "thesis", "synopsis_thesis_status", "synopsis/thesis status"],
        "imm_status": ["imm_status", "imm", "imm_status_remarks"],
    },
    "programs": {
        "program_code": ["program_code", "code"],
        "program_name": ["program_name", "name", "program"],
        "duration_months": ["duration_months", "duration", "months"],
    },
    "training_records": {
        "resident_email": ["resident_email", "email"],
        "program_code": ["program_code", "program", "track"],
        "start_date": ["start_date", "training_start", "date_joining"],
        "expected_end_date": ["expected_end_date", "training_end", "end_date"],
        "current_level": ["current_level", "training_level", "level"],
    },
}

SPECIALTY_BY_DEPARTMENT = {
    "MED": "medicine",
    "SURG": "surgery",
    "PED": "pediatrics",
    "OBG": "gynecology",
    "ORTH": "orthopedics",
}


@dataclass
class ImportArtifacts:
    supervisors_rows: list[dict] = field(default_factory=list)
    residents_rows: list[dict] = field(default_factory=list)
    supervision_rows: list[dict] = field(default_factory=list)
    program_rows: list[dict] = field(default_factory=list)
    training_record_rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    discovered_files: dict[str, str] = field(default_factory=dict)


class Command(BaseCommand):
    help = (
        "Normalize a pilot workbook / resident list / supervisor list into the active PGSIMS "
        "bulk-import schema. Defaults to dry-run."
    )

    def add_arguments(self, parser):
        parser.add_argument("--apply", action="store_true", help="Write changes to the database.")
        parser.add_argument("--workbook", type=str, help="Path to the finalized pilot workbook.")
        parser.add_argument("--supervisors-file", type=str, help="Path to the finalized supervisors list.")
        parser.add_argument("--residents-file", type=str, help="Path to the finalized residents list.")
        parser.add_argument(
            "--admin-username",
            type=str,
            default="admin",
            help="Admin username used as the actor for import operations.",
        )

    def handle(self, *args, **options):
        actor = User.objects.filter(username=options["admin_username"], role="admin").first()
        if not actor:
            raise CommandError(f"Admin user '{options['admin_username']}' not found.")

        artifacts = self._build_artifacts(
            workbook_path=options.get("workbook"),
            supervisors_path=options.get("supervisors_file"),
            residents_path=options.get("residents_file"),
        )

        self.stdout.write(self.style.WARNING("Pilot import bundle"))
        self.stdout.write(f"Mode: {'APPLY' if options['apply'] else 'DRY-RUN'}")
        for label, path in sorted(artifacts.discovered_files.items()):
            self.stdout.write(f"{label}: {path}")
        self.stdout.write(f"supervisors_rows: {len(artifacts.supervisors_rows)}")
        self.stdout.write(f"residents_rows: {len(artifacts.residents_rows)}")
        self.stdout.write(f"supervision_rows: {len(artifacts.supervision_rows)}")
        self.stdout.write(f"program_rows: {len(artifacts.program_rows)}")
        self.stdout.write(f"training_record_rows: {len(artifacts.training_record_rows)}")
        if artifacts.warnings:
            self.stdout.write("warnings:")
            for warning in artifacts.warnings:
                self.stdout.write(f"  - {warning}")

        service = BulkService(actor)
        dry_run = not options["apply"]
        results = {}

        if artifacts.program_rows:
            results["training-programs"] = self._run_bulk_service(
                service.import_training_programs,
                "training_programs.csv",
                artifacts.program_rows,
                dry_run=dry_run,
            )
        if artifacts.supervisors_rows:
            results["faculty-supervisors"] = self._run_userbase_import(
                actor,
                "faculty-supervisors",
                "supervisors.csv",
                artifacts.supervisors_rows,
                dry_run=dry_run,
            )
        if artifacts.residents_rows:
            results["residents"] = self._run_userbase_import(
                actor,
                "residents",
                "residents.csv",
                artifacts.residents_rows,
                dry_run=dry_run,
            )
        if artifacts.supervision_rows:
            results["supervision-links"] = self._run_userbase_import(
                actor,
                "supervision-links",
                "supervision_links.csv",
                artifacts.supervision_rows,
                dry_run=dry_run,
            )
        if artifacts.training_record_rows:
            results["resident-training-records"] = self._run_bulk_service(
                service.import_resident_training_records,
                "resident_training_records.csv",
                artifacts.training_record_rows,
                dry_run=dry_run,
            )

        for label, payload in results.items():
            successes = len(payload.get("successes", []))
            failures = len(payload.get("failures", []))
            self.stdout.write(f"{label}: successes={successes} failures={failures}")
            for failure in payload.get("failures", [])[:10]:
                self.stdout.write(f"  failure row {failure.get('row', '?')}: {failure.get('error')}")

        if not results:
            raise CommandError("No importable rows were produced from the supplied files.")

        self.stdout.write(self.style.SUCCESS("Pilot bundle normalization complete."))

    def _build_artifacts(
        self,
        *,
        workbook_path: str | None,
        supervisors_path: str | None,
        residents_path: str | None,
    ) -> ImportArtifacts:
        artifacts = ImportArtifacts()
        workbook = self._resolve_path(workbook_path, "workbook")
        supervisors = self._resolve_path(supervisors_path, "supervisors")
        residents = self._resolve_path(residents_path, "residents")

        if workbook:
            artifacts.discovered_files["workbook"] = str(workbook)
            sheet_rows = self._read_workbook(workbook)
            self._map_workbook(sheet_rows, artifacts)
        if supervisors:
            artifacts.discovered_files["supervisors"] = str(supervisors)
            rows = self._read_tabular_file(supervisors)
            artifacts.supervisors_rows.extend(self._map_supervisors(rows, artifacts))
        if residents:
            artifacts.discovered_files["residents"] = str(residents)
            rows = self._read_tabular_file(residents)
            artifacts.residents_rows.extend(self._map_residents(rows, artifacts))

        if not artifacts.discovered_files:
            raise CommandError(
                "No finalized pilot workbook or resident/supervisor lists were found. "
                "Supply --workbook/--supervisors-file/--residents-file explicitly."
            )

        supervisor_email_by_name = {
            row["full_name"].strip().lower(): row["email"]
            for row in artifacts.supervisors_rows
            if row.get("full_name") and row.get("email")
        }
        resident_email_by_name = {
            row["full_name"].strip().lower(): row["email"]
            for row in artifacts.residents_rows
            if row.get("full_name") and row.get("email")
        }

        seen_links = set()
        for resident_row in artifacts.residents_rows:
            supervisor_email = resident_row.pop("_supervisor_email", "") or ""
            supervisor_name = resident_row.pop("_supervisor_name", "") or ""
            if not supervisor_email and supervisor_name:
                supervisor_email = supervisor_email_by_name.get(supervisor_name.strip().lower(), "")
            if not supervisor_email:
                continue
            link_key = (supervisor_email.lower(), resident_row["email"].lower(), resident_row["department_code"])
            if link_key in seen_links:
                continue
            seen_links.add(link_key)
            artifacts.supervision_rows.append(
                {
                    "supervisor_email": supervisor_email,
                    "resident_email": resident_row["email"],
                    "department_code": resident_row.get("department_code", ""),
                    "start_date": resident_row["training_start"],
                    "end_date": "",
                    "active": "true",
                }
            )

        explicit_training_rows = {tuple(sorted(row.items())) for row in artifacts.training_record_rows}
        if not artifacts.program_rows:
            artifacts.warnings.append(
                "No explicit training program sheet/columns were found. Program and training-record imports were skipped."
            )
        else:
            if not artifacts.training_record_rows:
                for resident_row in artifacts.residents_rows:
                    track_value = (resident_row.pop("_program_track", "") or "").strip()
                    if not track_value:
                        continue
                    program_code = self._normalize_program_code(track_value)
                    if not program_code:
                        continue
                    training_row = {
                        "resident_email": resident_row["email"],
                        "program_code": program_code,
                        "start_date": resident_row["training_start"],
                        "expected_end_date": resident_row.get("training_end", ""),
                        "current_level": resident_row.get("training_level", ""),
                        "active": "true",
                    }
                    key = tuple(sorted(training_row.items()))
                    if key not in explicit_training_rows:
                        explicit_training_rows.add(key)
                        artifacts.training_record_rows.append(training_row)

        unresolved_status_columns = [
            row["full_name"]
            for row in artifacts.residents_rows
            if row.pop("_thesis_status", "") or row.pop("_imm_status", "")
        ]
        if unresolved_status_columns:
            artifacts.warnings.append(
                "Thesis/IMM columns were detected but not imported automatically because the "
                "active schema has no canonical bulk surface for those placeholders."
            )

        # Deduplicate identical rows.
        artifacts.supervisors_rows = self._dedupe(artifacts.supervisors_rows)
        artifacts.residents_rows = self._dedupe(artifacts.residents_rows)
        artifacts.supervision_rows = self._dedupe(artifacts.supervision_rows)
        artifacts.program_rows = self._dedupe(artifacts.program_rows)
        artifacts.training_record_rows = self._dedupe(artifacts.training_record_rows)
        return artifacts

    def _resolve_path(self, raw_path: str | None, kind: str) -> Path | None:
        if raw_path:
            path = Path(raw_path).expanduser().resolve()
            if not path.exists():
                raise CommandError(f"{kind} file not found: {path}")
            return path
        matches: list[Path] = []
        for root in SEARCH_ROOTS:
            if not root.exists():
                continue
            for pattern in DISCOVERY_PATTERNS[kind]:
                matches.extend(
                    path
                    for path in root.rglob(pattern)
                    if path.is_file()
                    and path.suffix.lower() in SPREADSHEET_SUFFIXES
                    and "frontend/public/templates" not in str(path)
                    and "PGSIMS_Demo_CaseSeed" not in path.name
                    and "template" not in path.name.lower()
                    and "demo" not in path.name.lower()
                )
        if not matches:
            return None
        return sorted(matches, key=lambda item: item.stat().st_mtime, reverse=True)[0]

    def _read_workbook(self, path: Path) -> dict[str, list[dict]]:
        workbook = load_workbook(path, data_only=True)
        payload: dict[str, list[dict]] = {}
        for sheet in workbook.worksheets:
            headers = [
                self._normalize_header(cell.value)
                for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=False))
            ]
            rows: list[dict] = []
            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                row = {
                    headers[index]: str(value).strip() if value is not None else ""
                    for index, value in enumerate(values)
                    if index < len(headers) and headers[index]
                }
                row["_row_number"] = row_number
                rows.append(row)
            if rows:
                payload[sheet.title] = rows
        return payload

    def _read_tabular_file(self, path: Path) -> list[dict]:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                return [
                    {
                        **{self._normalize_header(key): (value or "").strip() for key, value in row.items() if key},
                        "_row_number": index,
                    }
                    for index, row in enumerate(reader, start=2)
                    if any((value or "").strip() for value in row.values())
                ]
        return next(iter(self._read_workbook(path).values()), [])

    def _map_workbook(self, sheet_rows: dict[str, list[dict]], artifacts: ImportArtifacts) -> None:
        for sheet_name, rows in sheet_rows.items():
            lowered = sheet_name.lower()
            if any(hint in lowered for hint in SHEET_HINTS["supervisors"]):
                artifacts.supervisors_rows.extend(self._map_supervisors(rows, artifacts))
            elif any(hint in lowered for hint in SHEET_HINTS["residents"]):
                artifacts.residents_rows.extend(self._map_residents(rows, artifacts))
            elif any(hint in lowered for hint in SHEET_HINTS["programs"]):
                artifacts.program_rows.extend(self._map_programs(rows, artifacts))
            elif any(hint in lowered for hint in SHEET_HINTS["training_records"]):
                artifacts.training_record_rows.extend(self._map_training_records(rows, artifacts))

    def _map_supervisors(self, rows: list[dict], artifacts: ImportArtifacts) -> list[dict]:
        department_map = self._department_code_map()
        default_hospital = self._default_hospital_code()
        used_usernames = set(User.objects.values_list("username", flat=True))
        normalized_rows = []
        for row in rows:
            full_name = self._pick(row, ALIASES["supervisors"]["full_name"])
            if not full_name:
                continue
            department_code = self._resolve_department_code(
                self._pick(row, ALIASES["supervisors"]["department_code"]),
                department_map,
            )
            specialty = self._pick(row, ALIASES["supervisors"]["specialty"]) or SPECIALTY_BY_DEPARTMENT.get(
                department_code,
                "other",
            )
            username = self._generate_unique_username(full_name, used_usernames)
            normalized_rows.append(
                {
                    "email": self._pick(row, ALIASES["supervisors"]["email"]) or self._placeholder_email(username, "supervisor"),
                    "full_name": full_name,
                    "phone_number": self._pick(row, ALIASES["supervisors"]["phone_number"]),
                    "role": self._pick(row, ALIASES["supervisors"]["role"]) or "supervisor",
                    "specialty": specialty,
                    "department_code": department_code,
                    "hospital_code": self._resolve_hospital_code(
                        self._pick(row, ALIASES["supervisors"]["hospital_code"]),
                        default_hospital,
                    ),
                    "designation": self._pick(row, ALIASES["supervisors"]["designation"]),
                    "registration_number": self._pick(row, ALIASES["supervisors"]["registration_number"]),
                    "username": username,
                    "password": "",
                    "active": "true",
                    "start_date": self._normalize_date(
                        self._pick(row, ALIASES["supervisors"]["start_date"]) or date.today().isoformat(),
                        artifacts,
                        context=f"supervisor:{full_name}",
                    ),
                }
            )
        return normalized_rows

    def _map_residents(self, rows: list[dict], artifacts: ImportArtifacts) -> list[dict]:
        department_map = self._department_code_map()
        default_hospital = self._default_hospital_code()
        used_usernames = set(User.objects.values_list("username", flat=True))
        normalized_rows = []
        for row in rows:
            full_name = self._pick(row, ALIASES["residents"]["full_name"])
            if not full_name:
                continue
            department_code = self._resolve_department_code(
                self._pick(row, ALIASES["residents"]["department_code"]),
                department_map,
            )
            specialty = self._pick(row, ALIASES["residents"]["specialty"]) or SPECIALTY_BY_DEPARTMENT.get(
                department_code,
                "other",
            )
            username = self._generate_unique_username(full_name, used_usernames)
            training_start_raw = self._pick(row, ALIASES["residents"]["training_start"]) or date.today().isoformat()
            training_start = self._normalize_date(training_start_raw, artifacts, context=f"resident:{full_name}")
            year_value = self._normalize_year(self._pick(row, ALIASES["residents"]["year"]), training_start)
            normalized_rows.append(
                {
                    "email": self._pick(row, ALIASES["residents"]["email"]) or self._placeholder_email(username, "resident"),
                    "full_name": full_name,
                    "phone_number": self._pick(row, ALIASES["residents"]["phone_number"]),
                    "role": self._pick(row, ALIASES["residents"]["role"]) or "resident",
                    "specialty": specialty,
                    "year": year_value,
                    "pgr_id": self._pick(row, ALIASES["residents"]["pgr_id"]),
                    "training_start": training_start,
                    "training_end": self._normalize_optional_date(
                        self._pick(row, ALIASES["residents"]["training_end"]),
                        artifacts,
                        context=f"resident:{full_name}:training_end",
                    ),
                    "training_level": self._pick(row, ALIASES["residents"]["training_level"]) or f"y{year_value}",
                    "department_code": department_code,
                    "hospital_code": self._resolve_hospital_code(
                        self._pick(row, ALIASES["residents"]["hospital_code"]),
                        default_hospital,
                    ),
                    "supervisor_email": "",
                    "username": username,
                    "password": "",
                    "active": "true",
                    "_supervisor_email": self._pick(row, ALIASES["residents"]["supervisor_email"]),
                    "_supervisor_name": self._pick(row, ALIASES["residents"]["supervisor_name"]),
                    "_program_track": self._pick(row, ALIASES["residents"]["program_track"]),
                    "_thesis_status": self._pick(row, ALIASES["residents"]["thesis_status"]),
                    "_imm_status": self._pick(row, ALIASES["residents"]["imm_status"]),
                }
            )
        return normalized_rows

    def _map_programs(self, rows: list[dict], artifacts: ImportArtifacts) -> list[dict]:
        normalized_rows = []
        for row in rows:
            code = self._pick(row, ALIASES["programs"]["program_code"])
            name = self._pick(row, ALIASES["programs"]["program_name"])
            duration = self._pick(row, ALIASES["programs"]["duration_months"])
            if not code or not name or not duration:
                continue
            normalized_rows.append(
                {
                    "program_code": self._normalize_program_code(code),
                    "program_name": name,
                    "duration_months": duration,
                    "active": "true",
                }
            )
        return normalized_rows

    def _map_training_records(self, rows: list[dict], artifacts: ImportArtifacts) -> list[dict]:
        normalized_rows = []
        for row in rows:
            resident_email = self._pick(row, ALIASES["training_records"]["resident_email"])
            program_code = self._normalize_program_code(self._pick(row, ALIASES["training_records"]["program_code"]))
            start_date = self._normalize_optional_date(
                self._pick(row, ALIASES["training_records"]["start_date"]),
                artifacts,
                context=f"training_record:{resident_email or '?'}",
            )
            if not resident_email or not program_code or not start_date:
                continue
            normalized_rows.append(
                {
                    "resident_email": resident_email,
                    "program_code": program_code,
                    "start_date": start_date,
                    "expected_end_date": self._normalize_optional_date(
                        self._pick(row, ALIASES["training_records"]["expected_end_date"]),
                        artifacts,
                        context=f"training_record:{resident_email}:end",
                    ),
                    "current_level": self._pick(row, ALIASES["training_records"]["current_level"]),
                    "active": "true",
                }
            )
        return normalized_rows

    def _run_userbase_import(self, actor, entity: str, filename: str, rows: list[dict], *, dry_run: bool) -> dict:
        upload = SimpleUploadedFile(filename, self._rows_to_csv_bytes(rows), content_type="text/csv")
        return import_entity(actor, entity, upload, dry_run=dry_run, allow_partial=False)

    def _run_bulk_service(self, callback, filename: str, rows: list[dict], *, dry_run: bool) -> dict:
        upload = SimpleUploadedFile(filename, self._rows_to_csv_bytes(rows), content_type="text/csv")
        operation = callback(upload, dry_run=dry_run, allow_partial=False)
        return operation.details

    def _rows_to_csv_bytes(self, rows: list[dict]) -> bytes:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")

    def _pick(self, row: dict, aliases: list[str]) -> str:
        for alias in aliases:
            normalized = self._normalize_header(alias)
            if normalized in row and row[normalized]:
                return str(row[normalized]).strip()
        return ""

    def _normalize_header(self, value) -> str:
        text = str(value or "").strip().lower()
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return text.strip("_")

    def _department_code_map(self) -> dict[str, str]:
        mapping = {}
        for department in Department.objects.filter(active=True):
            mapping[department.code.lower()] = department.code
            mapping[department.name.lower()] = department.code
        return mapping

    def _default_hospital_code(self) -> str:
        hospital = Hospital.objects.filter(is_active=True).order_by("id").first()
        return hospital.code if hospital and hospital.code else ""

    def _resolve_department_code(self, raw_value: str, department_map: dict[str, str]) -> str:
        value = (raw_value or "").strip()
        if not value:
            return ""
        return department_map.get(value.lower(), value.upper())

    def _resolve_hospital_code(self, raw_value: str, default_code: str) -> str:
        value = (raw_value or "").strip()
        return value.upper() if value else default_code

    def _generate_unique_username(self, full_name: str, used_usernames: set[str]) -> str:
        base = re.sub(r"[^a-z0-9]+", ".", full_name.lower()).strip(".") or "pilot.user"
        candidate = base
        suffix = 1
        while candidate in used_usernames:
            suffix += 1
            candidate = f"{base}.{suffix}"
        used_usernames.add(candidate)
        return candidate

    def _placeholder_email(self, username: str, role: str) -> str:
        return f"{username}.{role}@pilot-placeholder.local"

    def _normalize_date(self, raw_value: str, artifacts: ImportArtifacts, *, context: str) -> str:
        normalized = self._normalize_optional_date(raw_value, artifacts, context=context)
        if not normalized:
            fallback = date.today().isoformat()
            artifacts.warnings.append(f"{context}: missing/invalid date; defaulted to {fallback}.")
            return fallback
        return normalized

    def _normalize_optional_date(self, raw_value: str, artifacts: ImportArtifacts, *, context: str) -> str:
        raw = (raw_value or "").strip()
        if not raw:
            return ""
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt).date().isoformat()
            except Exception:
                continue
        artifacts.warnings.append(f"{context}: could not parse date '{raw}'.")
        return ""

    def _normalize_year(self, raw_value: str, training_start: str) -> str:
        value = (raw_value or "").strip().lower()
        if value in {"1", "2", "3", "4", "5"}:
            return value
        match = re.search(r"([1-5])", value)
        if match:
            return match.group(1)
        try:
            start_year = int(training_start[:4])
            delta = max(1, min(5, date.today().year - start_year + 1))
            return str(delta)
        except Exception:
            return "1"

    def _normalize_program_code(self, raw_value: str) -> str:
        value = (raw_value or "").strip()
        return re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").upper()

    def _dedupe(self, rows: list[dict]) -> list[dict]:
        seen = set()
        deduped = []
        for row in rows:
            key = tuple(sorted(row.items()))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        return deduped
