"""Service layer for bulk operations."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Iterable, Iterator, List, Optional, Sequence, Tuple

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from sims.bulk.models import BulkOperation
from sims.logbook.models import LogbookEntry
from sims.users.models import User


@dataclass
class BulkResult:
    successes: List[dict]
    failures: List[dict]


class BulkProcessingError(Exception):
    """Raised when a bulk operation fails in transactional mode."""


class BulkService:
    def __init__(self, actor: User, chunk_size: int = 50):
        self.actor = actor
        self.chunk_size = chunk_size
        self._validate_permissions()

    def _validate_permissions(self) -> None:
        if not (
            self.actor.is_superuser or getattr(self.actor, "role", None) in {"admin", "supervisor"}
        ):
            raise PermissionDenied("Bulk operations are restricted to supervisors and admins.")

    # ------------------------------------------------------------------
    # Review and assignment operations

    def review_entries(self, entry_ids: Sequence[int], status: str) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_REVIEW)
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.status = status
                        entry.supervisor_action_at = timezone.now()
                        entry.save(update_fields=["status", "supervisor_action_at"])
                        successes.append({"id": entry.pk, "status": status})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    def assign_supervisor(self, entry_ids: Sequence[int], supervisor: User) -> BulkOperation:
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_ASSIGNMENT
        )
        successes: List[dict] = []
        failures: List[dict] = []
        for chunk in _chunked(entry_ids, self.chunk_size):
            try:
                with transaction.atomic():
                    entries = list(LogbookEntry.objects.select_for_update().filter(pk__in=chunk))
                    missing = set(chunk) - {entry.pk for entry in entries}
                    for missing_id in missing:
                        failures.append({"id": missing_id, "error": "not-found"})
                    for entry in entries:
                        entry.supervisor = supervisor
                        entry.save(update_fields=["supervisor"])
                        successes.append({"id": entry.pk, "supervisor": supervisor.pk})
            except ValidationError as exc:
                failures.append({"ids": list(chunk), "error": str(exc)})
        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    # ------------------------------------------------------------------
    # Bulk import

    def import_logbook_entries(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
    ) -> BulkOperation:
        operation = BulkOperation.objects.create(user=self.actor, operation=BulkOperation.OP_IMPORT)
        rows = list(_parse_rows(uploaded_file))
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False

        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            try:
                pg = User.objects.get(username=row["pg_username"], role="pg")
            except User.DoesNotExist:
                failures.append({"row": row, "error": "invalid-pg"})
                errors_triggered = True
                return
            try:
                entry_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
            except (KeyError, ValueError):
                failures.append({"row": row, "error": "invalid-date"})
                errors_triggered = True
                return
            status = row.get("status", "draft")
            payload = {
                "pg": pg,
                "case_title": row.get("case_title") or "Untitled",
                "date": entry_date,
                "status": status,
                "location_of_activity": row.get("location") or "Not specified",
                "patient_history_summary": row.get("patient_history") or "Pending summary",
                "management_action": row.get("management_action") or "Pending action",
                "topic_subtopic": row.get("topic_subtopic") or "General",
            }
            try:
                if dry_run:
                    entry = LogbookEntry(**payload)
                    entry.full_clean()
                else:
                    entry = LogbookEntry.objects.create(**payload, created_by=self.actor)
                successes.append(
                    {
                        "pg": pg.username,
                        "case_title": payload["case_title"],
                        "status": status,
                    }
                )
            except ValidationError as exc:
                failures.append({"row": row, "error": exc.message_dict})
                errors_triggered = True

        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation

        operation.mark_completed(
            len(successes),
            len(failures),
            {"successes": successes, "failures": failures},
        )
        return operation

    def import_trainees(
        self,
        uploaded_file,
        *,
        dry_run: bool = True,
        allow_partial: bool = False,
    ) -> BulkOperation:
        """
        Import trainee data from Excel file.
        
        Expected columns:
        - Sr. No. (ignored)
        - Name of Trainee
        - Date of Joining
        - MS/FCPS (optional)
        - Supervisor Name (optional, will create if not found)
        """
        operation = BulkOperation.objects.create(
            user=self.actor, operation=BulkOperation.OP_IMPORT
        )
        rows = list(_parse_trainee_rows(uploaded_file))
        successes: List[dict] = []
        failures: List[dict] = []
        errors_triggered = False
        created_supervisors: List[dict] = []

        def process_row(row: dict) -> None:
            nonlocal errors_triggered
            
            row_num = row.get("_row_number", "unknown")
            name = row.get("name", "").strip()
            date_joining = row.get("date_joining")
            qualification = row.get("qualification", "").strip()
            supervisor_name = row.get("supervisor_name", "").strip()
            
            # Validate required fields
            if not name:
                failures.append({
                    "row": row_num,
                    "error": "Missing 'Name of Trainee'",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse name
            try:
                first_name, last_name = _parse_name(name)
                if not first_name and not last_name:
                    failures.append({
                        "row": row_num,
                        "error": "Invalid name format",
                        "data": row
                    })
                    errors_triggered = True
                    return
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error parsing name: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Parse date
            try:
                date_joined = _parse_date(str(date_joining)) if date_joining else date.today()
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Invalid date format: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Generate username
            try:
                username = _generate_username(first_name, last_name)
            except Exception as e:
                failures.append({
                    "row": row_num,
                    "error": f"Error generating username: {str(e)}",
                    "data": row
                })
                errors_triggered = True
                return
            
            # Infer training year
            try:
                year = _infer_training_year(date_joined)
            except Exception as e:
                year = "1"  # Default to year 1 on error
                failures.append({
                    "row": row_num,
                    "warning": f"Could not infer year, defaulting to 1: {str(e)}",
                    "data": row
                })
            
            # Get or create supervisor
            supervisor = None
            if supervisor_name:
                try:
                    supervisor = _get_or_create_supervisor(supervisor_name, self.actor)
                    if supervisor and supervisor.id not in [s["id"] for s in created_supervisors]:
                        created_supervisors.append({
                            "id": supervisor.id,
                            "username": supervisor.username,
                            "name": supervisor.get_full_name()
                        })
                except Exception as e:
                    failures.append({
                        "row": row_num,
                        "warning": f"Error with supervisor '{supervisor_name}': {str(e)}",
                        "data": row
                    })
                    # Continue without supervisor if allow_partial, otherwise fail
                    if not allow_partial:
                        errors_triggered = True
                        return
            else:
                # Supervisor name not provided
                if not allow_partial:
                    failures.append({
                        "row": row_num,
                        "error": "Supervisor Name is required",
                        "data": row
                    })
                    errors_triggered = True
                    return
            
            # Generate email
            email = f"{username}.pgr@pmc.edu.pk"
            
            # Prepare user data
            user_data = {
                "username": username,
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": "pg",
                "specialty": "urology",
                "year": year,
                "date_joined": date_joined,
                "registration_number": qualification if qualification else "",
                "is_active": True,
                "created_by": self.actor,
            }
            
            if supervisor:
                user_data["supervisor"] = supervisor
            
            # Create or validate user
            try:
                if dry_run:
                    # Validate without creating
                    user = User(**user_data)
                    user.full_clean()
                else:
                    # Check if user already exists
                    existing_user = User.objects.filter(username=username).first()
                    if existing_user:
                        # Update existing user
                        for key, value in user_data.items():
                            if key != "username" and key != "created_by":
                                setattr(existing_user, key, value)
                        existing_user.modified_by = self.actor
                        existing_user.full_clean()
                        existing_user.save()
                        user = existing_user
                    else:
                        # Create new user
                        user = User.objects.create(**user_data)
                        # Set password (default password, should be changed on first login)
                        user.set_password("changeme123")
                        user.save()
                
                successes.append({
                    "row": row_num,
                    "username": username,
                    "name": f"{first_name} {last_name}".strip(),
                    "email": email,
                    "year": year,
                    "supervisor": supervisor.username if supervisor else None,
                })
            except ValidationError as exc:
                failures.append({
                    "row": row_num,
                    "error": exc.message_dict if hasattr(exc, "message_dict") else str(exc),
                    "data": row
                })
                errors_triggered = True
            except Exception as exc:
                failures.append({
                    "row": row_num,
                    "error": f"Unexpected error: {str(exc)}",
                    "data": row
                })
                errors_triggered = True

        # Process rows
        if dry_run or allow_partial:
            for row in rows:
                process_row(row)
        else:
            try:
                with transaction.atomic():
                    for row in rows:
                        process_row(row)
                    if errors_triggered:
                        raise BulkProcessingError("Import failed; rolling back")
            except BulkProcessingError:
                operation.mark_failed({"failures": failures})
                return operation

        # Prepare operation details
        details = {
            "successes": successes,
            "failures": failures,
            "created_supervisors": created_supervisors,
        }
        
        operation.mark_completed(
            len(successes),
            len(failures),
            details,
        )
        return operation


def _chunked(items: Sequence[int], chunk_size: int) -> Iterator[List[int]]:
    chunk: List[int] = []
    for item in items:
        chunk.append(item)
        if len(chunk) == chunk_size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


REQUIRED_COLUMNS = {"pg_username", "case_title", "date", "status"}


def _parse_rows(uploaded_file) -> Iterator[dict]:
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    if name.endswith(".csv"):
        text_stream = (
            io.TextIOWrapper(stream, encoding="utf-8") if isinstance(stream, io.BytesIO) else stream
        )
        reader = csv.DictReader(text_stream)
        _validate_headers(reader.fieldnames)
        for row in reader:
            yield {key: (value or "").strip() for key, value in row.items() if key}
    elif name.endswith(".xlsx"):
        workbook = load_workbook(stream)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        _validate_headers(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            payload = {
                headers[idx]: (value or "") for idx, value in enumerate(row) if idx < len(headers)
            }
            yield {key: str(value).strip() for key, value in payload.items()}
    else:
        raise ValidationError("Unsupported file format")


def _validate_headers(headers: Iterable[str]) -> None:
    if not headers:
        raise ValidationError("No headers found in file")
    missing = REQUIRED_COLUMNS - set(header.strip() for header in headers if header)
    if missing:
        raise ValidationError(f"Missing columns: {', '.join(sorted(missing))}")


# ------------------------------------------------------------------
# Trainee import helper functions


def _parse_name(full_name: str) -> Tuple[str, str]:
    """
    Split full name into first_name and last_name.
    - Remove titles like Dr., Mr., Mrs., etc.
    - Last word → last_name
    - All other words → first_name
    """
    if not full_name or not full_name.strip():
        return ("", "")
    
    # Remove common titles
    title_pattern = r"^(Dr\.?|Mr\.?|Mrs\.?|Ms\.?|Prof\.?|Professor)\s+"
    name = re.sub(title_pattern, "", full_name.strip(), flags=re.IGNORECASE)
    
    # Split by spaces
    parts = [p.strip() for p in name.split() if p.strip()]
    
    if not parts:
        return ("", "")
    elif len(parts) == 1:
        return (parts[0], "")
    else:
        # Last word is last_name, rest is first_name
        return (" ".join(parts[:-1]), parts[-1])


def _generate_username(first_name: str, last_name: str) -> str:
    """
    Generate username as firstname.lastname.
    - Convert to lowercase
    - Remove special characters, keep only alphanumeric and dots
    - Check for duplicates and append number if needed
    """
    # Clean names
    first_clean = re.sub(r"[^a-z0-9]", "", first_name.lower())
    last_clean = re.sub(r"[^a-z0-9]", "", last_name.lower())
    
    if not first_clean and not last_clean:
        base_username = "trainee"
    elif not last_clean:
        base_username = first_clean
    elif not first_clean:
        base_username = last_clean
    else:
        base_username = f"{first_clean}.{last_clean}"
    
    # Check for duplicates
    username = base_username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}{counter}"
        counter += 1
    
    return username


def _infer_training_year(date_joined: date) -> str:
    """
    Calculate training year based on months since date_joined.
    - 0-12 months → "1"
    - 12-24 months → "2"
    - 24-36 months → "3"
    - 36+ months → "4"
    """
    today = date.today()
    if date_joined > today:
        # Future date, default to year 1
        return "1"
    
    # Calculate months difference
    months_diff = (today.year - date_joined.year) * 12 + (today.month - date_joined.month)
    
    if months_diff < 12:
        return "1"
    elif months_diff < 24:
        return "2"
    elif months_diff < 36:
        return "3"
    else:
        return "4"


def _get_or_create_supervisor(supervisor_name: str, actor: User) -> Optional[User]:
    """
    Find supervisor by name or create if not found.
    - Search by full name (case-insensitive)
    - If not found, create new supervisor user
    - Set specialty to urology
    - Generate username and email
    """
    if not supervisor_name or not supervisor_name.strip():
        return None
    
    supervisor_name = supervisor_name.strip()
    
    # Try to find existing supervisor by name
    # Search by first_name + last_name or full name match
    first_name, last_name = _parse_name(supervisor_name)
    
    if first_name and last_name:
        # Try exact match
        supervisor = User.objects.filter(
            first_name__iexact=first_name,
            last_name__iexact=last_name,
            role="supervisor"
        ).first()
        
        if supervisor:
            return supervisor
        
        # Try full name match
        supervisor = User.objects.filter(
            first_name__iexact=supervisor_name,
            role="supervisor"
        ).first()
        
        if supervisor:
            return supervisor
    
    # Create new supervisor
    username = _generate_username(first_name, last_name)
    email = f"{username}.supervisor@pmc.edu.pk"
    
    supervisor = User.objects.create(
        username=username,
        email=email,
        first_name=first_name or supervisor_name,
        last_name=last_name,
        role="supervisor",
        specialty="urology",
        is_active=True,
        created_by=actor,
    )
    
    return supervisor


def _parse_date(date_str: str) -> date:
    """
    Parse date string in various formats.
    Supports: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY, etc.
    """
    if not date_str:
        raise ValueError("Empty date string")
    
    date_str = str(date_str).strip()
    
    # Try different date formats
    formats = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%Y.%m.%d",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    # If all formats fail, try parsing as Excel date serial number
    try:
        # Excel dates are sometimes stored as numbers
        serial = float(date_str)
        # Excel epoch is 1900-01-01, but Excel incorrectly treats 1900 as a leap year
        excel_epoch = datetime(1899, 12, 30)
        parsed_date = excel_epoch + timedelta(days=int(serial))
        return parsed_date.date()
    except (ValueError, OverflowError):
        pass
    
    raise ValueError(f"Unable to parse date: {date_str}")


def _parse_trainee_rows(uploaded_file) -> Iterator[dict]:
    """
    Parse Excel file with trainee data.
    Expected columns: Sr. No., Name of Trainee, Date of Joining, MS/FCPS, Supervisor Name
    """
    name = getattr(uploaded_file, "name", "uploaded")
    content = uploaded_file.read()
    if isinstance(content, bytes):
        stream = io.BytesIO(content)
    else:
        stream = io.StringIO(content)
    stream.seek(0)
    
    if not name.endswith((".xlsx", ".xls")):
        raise ValidationError("Trainee import only supports Excel files (.xlsx, .xls)")
    
    workbook = load_workbook(stream)
    sheet = workbook.active
    
    # Get headers from first row
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(max_row=1))]
    
    # Map column indices
    # Expected: Sr. No., Name of Trainee, Date of Joining, MS/FCPS, Supervisor Name
    col_map = {}
    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if "sr" in header_lower and "no" in header_lower:
            col_map["sr_no"] = idx
        elif "name" in header_lower and "trainee" in header_lower:
            col_map["name"] = idx
        elif "date" in header_lower and "joining" in header_lower:
            col_map["date"] = idx
        elif "ms" in header_lower or "fcps" in header_lower:
            col_map["qualification"] = idx
        elif "supervisor" in header_lower and "name" in header_lower:
            col_map["supervisor"] = idx
    
    # Validate required columns
    if "name" not in col_map or "date" not in col_map:
        raise ValidationError("Missing required columns: 'Name of Trainee' and 'Date of Joining'")
    
    # Parse data rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue
        
        row_data = {}
        if "name" in col_map and col_map["name"] < len(row):
            row_data["name"] = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
        if "date" in col_map and col_map["date"] < len(row):
            row_data["date_joining"] = row[col_map["date"]]
        if "qualification" in col_map and col_map["qualification"] < len(row):
            row_data["qualification"] = str(row[col_map["qualification"]]).strip() if row[col_map["qualification"]] else ""
        if "supervisor" in col_map and col_map["supervisor"] < len(row):
            row_data["supervisor_name"] = str(row[col_map["supervisor"]]).strip() if row[col_map["supervisor"]] else ""
        
        row_data["_row_number"] = row_idx
        yield row_data
