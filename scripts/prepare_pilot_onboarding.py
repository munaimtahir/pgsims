#!/usr/bin/env python3
import openpyxl
import csv
import re
from datetime import datetime, date

# Mapping of supervisor names to emails
SUPERVISOR_MAP = {
    "prof. dr. amir shaukat": "amir.shaukat@placeholder.example.com",
    "prof aamir shaukat sb": "amir.shaukat@placeholder.example.com",
    "prof dr.aamir shoukat": "amir.shaukat@placeholder.example.com",
    "dr. muhammad irfan": "muhammad.irfan@placeholder.example.com",
    "dr m": "muhammad.irfan@placeholder.example.com",
    "dr. umair ahmed": "umair.ahmed@placeholder.example.com",
    "dr.umair ahmed": "umair.ahmed@placeholder.example.com",
    "dr. muhammad irfan munir": "muhammad.irfan.munir@placeholder.example.com",
    "dr. muhammad irfan munir sb": "muhammad.irfan.munir@placeholder.example.com",
    "dr.muhammad irfan munir": "muhammad.irfan.munir@placeholder.example.com",
    "prof. dr. m. tahir bashir malik": "dr.m.tahir.bashir.malik@placeholder.example.com",
    "prof. dr. muhammad akmal": "dr.muhammad.akmal@placeholder.example.com",
    "dr. muhammad akmal": "dr.muhammad.akmal@placeholder.example.com",
    "dr. sheraz javed": "sheraz.javed@placeholder.example.com",
    "dr muhammad arif": "muhammad.arif@placeholder.example.com",
    "dr shahid abbas": "shahid.abbas@placeholder.example.com",
}

# Master list of supervisors to generate
SUPERVISORS_DATA = [
    # Urology
    {"email": "dr.muhammad.akmal@placeholder.example.com", "full_name": "Prof. Dr. Muhammad Akmal", "role": "supervisor", "specialty": "urology", "department_code": "SURG", "hospital_code": "UTRMC", "designation": "Professor"},
    {"email": "muhammad.irfan.munir@placeholder.example.com", "full_name": "Dr. Muhammad Irfan Munir", "role": "supervisor", "specialty": "urology", "department_code": "SURG", "hospital_code": "UTRMC", "designation": "Associate Professor"},
    {"email": "dr.m.tahir.bashir.malik@placeholder.example.com", "full_name": "Prof. Dr. M. Tahir Bashir Malik", "role": "supervisor", "specialty": "urology", "department_code": "SURG", "hospital_code": "UTRMC", "designation": "Professor"},
    {"email": "sheraz.javed@placeholder.example.com", "full_name": "Dr. Sheraz Javed", "role": "supervisor", "specialty": "urology", "department_code": "SURG", "hospital_code": "UTRMC", "designation": "Assistant Professor"},
    # Medicine & Allied
    {"email": "amir.shaukat@placeholder.example.com", "full_name": "Prof. Dr. Amir Shaukat", "role": "supervisor", "specialty": "medicine", "department_code": "MED", "hospital_code": "AH", "designation": "Professor"},
    {"email": "muhammad.irfan@placeholder.example.com", "full_name": "Dr. Muhammad Irfan", "role": "supervisor", "specialty": "medicine", "department_code": "MED", "hospital_code": "AH", "designation": "Associate Professor"},
    {"email": "umair.ahmed@placeholder.example.com", "full_name": "Dr. Umair Ahmed", "role": "supervisor", "specialty": "medicine", "department_code": "MED", "hospital_code": "AH", "designation": "Assistant Professor"},
    {"email": "muhammad.arif@placeholder.example.com", "full_name": "Dr. Muhammad Arif", "role": "supervisor", "specialty": "medicine", "department_code": "MED", "hospital_code": "AH", "designation": "Assistant Professor"},
    {"email": "shahid.abbas@placeholder.example.com", "full_name": "Dr. Shahid Abbas", "role": "supervisor", "specialty": "medicine", "department_code": "MED", "hospital_code": "AH", "designation": "Assistant Professor"},
]

def clean_val(v):
    if v is None: return ""
    return str(v).strip()

def clean_phone(v):
    val = clean_val(v)
    if not val: return ""
    # strip .0 if float
    if val.endswith(".0"):
        val = val[:-2]
    # replace 92 with 0 for consistency
    if val.startswith("92"):
        val = "0" + val[2:]
    return val

def generate_username(full_name, used):
    name = clean_val(full_name).lower()
    name = re.sub(r"^(dr|prof|syed)\.?\s+", "", name)  # strip titles
    base = re.sub(r"[^a-z0-9]+", ".", name).strip(".")
    if not base:
        base = "resident"
    candidate = base
    suffix = 1
    while candidate in used:
        suffix += 1
        candidate = f"{base}.{suffix}"
    used.add(candidate)
    return candidate

def parse_date(date_val):
    if not date_val:
        return ""
    if isinstance(date_val, (datetime, date)):
        return date_val.strftime("%Y-%m-%d")
    raw = str(date_val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # try parsing just date part if there is space
    if " " in raw:
        try:
            return datetime.strptime(raw.split(" ")[0], "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return raw

def resolve_specialty(program, allied_spec):
    prog = clean_val(program).lower()
    allied = clean_val(allied_spec).lower()
    
    if "urology" in prog:
        return "urology", "SURG"
    
    if "oncology" in allied or "oncology" in prog:
        return "other", "ONCO"
    if "cardiology" in allied or "cardiology" in prog:
        return "cardiology", "CARD"
    
    return "medicine", "MED"

def resolve_hospital(posting):
    p = clean_val(posting).lower()
    if "fic" in p or "cardiology" in p:
        return "AH"  # Default allied/FIC is mapped to AH or we can use UTRMC. Let's use AH.
    if "allied" in p or "ahf" in p or "mu 1" in p or "medical unit" in p:
        return "AH"
    return "UTRMC"

def main():
    used_usernames = {"admin", "e2e_admin", "pilot_admin"}
    residents = []
    
    # 1. Process PGSIMS Pilot Resident Responses (1).xlsx
    wb1 = openpyxl.load_workbook("PGSIMS Pilot Resident Responses (1).xlsx", data_only=True)
    sheet1 = wb1["Form Responses 1"]
    rows1 = list(sheet1.iter_rows(values_only=True))
    
    print(f"Loaded {len(rows1)-1} rows from File 1")
    
    for idx, row in enumerate(rows1[1:], start=2):
        name = clean_val(row[2])
        if not name:
            continue
        email_base = clean_val(row[1])
        email_gmail = clean_val(row[9])
        email = email_gmail if email_gmail else email_base
        phone = clean_phone(row[8])
        program = clean_val(row[15])
        allied_spec = clean_val(row[16])
        year_str = clean_val(row[17])
        # extract digit from 'Year 3' or '3'
        year_match = re.search(r"([1-5])", year_str)
        year = year_match.group(1) if year_match else "1"
        
        induction = parse_date(row[18])
        expected_end = parse_date(row[19])
        reg_id = clean_val(row[23])
        supervisor_name = clean_val(row[24])
        posting = clean_val(row[33])
        
        specialty, dept_code = resolve_specialty(program, allied_spec)
        hospital_code = resolve_hospital(posting)
        
        # supervisor email resolution
        sup_email = SUPERVISOR_MAP.get(supervisor_name.lower())
        if not sup_email:
            # check primary supervisor name column 26
            primary_sup = clean_val(row[26]) if len(row) > 26 else None
            if primary_sup:
                sup_email = SUPERVISOR_MAP.get(primary_sup.lower())
        if not sup_email and len(row) > 35:
            # check research supervisor name column 35
            res_sup = clean_val(row[35])
            if res_sup:
                sup_email = SUPERVISOR_MAP.get(res_sup.lower())
        if not sup_email and len(row) > 51:
            # check co-supervisor/primary supervisor column 51
            co_sup = clean_val(row[51])
            if co_sup:
                sup_email = SUPERVISOR_MAP.get(co_sup.lower())
        if not sup_email:
            sup_email = "amir.shaukat@placeholder.example.com"  # default fallback
            
        username = generate_username(name, used_usernames)
        
        # If PGR ID is missing, assign a temporary one
        pgr_id = reg_id
        is_pgr_id_missing = False
        if not pgr_id or pgr_id.lower() == "none" or pgr_id.lower() == "nil":
            is_pgr_id_missing = True
            pgr_id = f"PGR-TEMP-{idx-1:03d}"
            
        residents.append({
            "email": email,
            "full_name": name,
            "phone_number": phone,
            "role": "resident",
            "specialty": specialty,
            "year": year,
            "pgr_id": pgr_id,
            "training_start": induction,
            "training_end": expected_end,
            "training_level": f"Y{year}",
            "department_code": dept_code,
            "hospital_code": hospital_code,
            "supervisor_email": sup_email,
            "username": username,
            "password": "",
            "active": "true",
            "_source_file": "PGSIMS Pilot Resident Responses (1).xlsx",
            "_source_row": idx,
            "_pgr_id_missing": is_pgr_id_missing,
            "_original_supervisor": supervisor_name,
        })
        
    # 2. Process UROLOGY PGSIMS Pilot Resident Data Form (Responses).xlsx
    wb2 = openpyxl.load_workbook("UROLOGY PGSIMS Pilot Resident Data Form (Responses).xlsx", data_only=True)
    sheet2 = wb2["Form Responses 1"]
    rows2 = list(sheet2.iter_rows(values_only=True))
    
    print(f"Loaded {len(rows2)-1} rows from File 2")
    
    for idx, row in enumerate(rows2[1:], start=2):
        name = clean_val(row[2])
        if not name:
            continue
        email_base = clean_val(row[1])
        email_gmail = clean_val(row[8])
        email = email_gmail if email_gmail else email_base
        phone = clean_phone(row[7])
        program = clean_val(row[14])
        year_str = clean_val(row[15])
        year_match = re.search(r"([1-5])", year_str)
        year = year_match.group(1) if year_match else "1"
        
        induction = parse_date(row[16])
        expected_end = parse_date(row[17])
        reg_id = clean_val(row[19])
        supervisor_name = clean_val(row[20])
        posting = clean_val(row[24])
        
        specialty, dept_code = "urology", "SURG"
        hospital_code = resolve_hospital(posting)
        
        sup_email = SUPERVISOR_MAP.get(supervisor_name.lower())
        if not sup_email:
            sup_email = "muhammad.irfan.munir@placeholder.example.com"
            
        username = generate_username(name, used_usernames)
        
        pgr_id = reg_id
        is_pgr_id_missing = False
        if not pgr_id or pgr_id.lower() == "none" or pgr_id.lower() == "nil":
            is_pgr_id_missing = True
            pgr_id = f"PGR-TEMP-URO-{idx-1:03d}"
            
        residents.append({
            "email": email,
            "full_name": name,
            "phone_number": phone,
            "role": "resident",
            "specialty": specialty,
            "year": year,
            "pgr_id": pgr_id,
            "training_start": induction,
            "training_end": expected_end,
            "training_level": f"Y{year}",
            "department_code": dept_code,
            "hospital_code": hospital_code,
            "supervisor_email": sup_email,
            "username": username,
            "password": "",
            "active": "true",
            "_source_file": "UROLOGY PGSIMS Pilot Resident Data Form (Responses).xlsx",
            "_source_row": idx,
            "_pgr_id_missing": is_pgr_id_missing,
            "_original_supervisor": supervisor_name,
        })
        
    print(f"\nProcessed {len(residents)} total residents.")
    
    # 3. Write clean residents CSV
    with open("pilot_data/first_pilot_run/residents_import.csv", "w", newline="") as f:
        fieldnames = ["email", "full_name", "phone_number", "role", "specialty", "year", "pgr_id", "training_start", "training_end", "training_level", "department_code", "hospital_code", "supervisor_email", "username", "password", "active"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in residents:
            # write only valid fields for import
            row_to_write = {k: r[k] for k in fieldnames}
            writer.writerow(row_to_write)
            
    print("Wrote pilot_data/first_pilot_run/residents_import.csv")
    
    # 4. Write clean supervisors CSV
    with open("pilot_data/first_pilot_run/supervisors_import.csv", "w", newline="") as f:
        fieldnames = ["email", "full_name", "phone_number", "role", "specialty", "department_code", "hospital_code", "designation", "registration_number", "username", "password", "active", "start_date"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        used_sup_usernames = set()
        for s in SUPERVISORS_DATA:
            username = generate_username(s["full_name"], used_sup_usernames)
            row_to_write = {
                "email": s["email"],
                "full_name": s["full_name"],
                "phone_number": "03001234567",
                "role": s["role"],
                "specialty": s["specialty"],
                "department_code": s["department_code"],
                "hospital_code": s["hospital_code"],
                "designation": s["designation"],
                "registration_number": f"PMC-SUP-{s['email'].split('@')[0].replace('.', '-')}",
                "username": username,
                "password": "",
                "active": "true",
                "start_date": "2026-01-01",
            }
            writer.writerow(row_to_write)
            
    print("Wrote pilot_data/first_pilot_run/supervisors_import.csv")
    
    # Print warnings for missing PGR IDs
    print("\n=== DATA VERIFICATION WARNINGS ===")
    for r in residents:
        if r["_pgr_id_missing"]:
            print(f"Warning: Missing Registration ID (PGR ID) for '{r['full_name']}' in row {r['_source_row']} of {r['_source_file']}. Assigned placeholder: {r['pgr_id']}.")

if __name__ == "__main__":
    main()
